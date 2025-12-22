#!/usr/bin/env python3
"""
Coupang Coupon Issuer - 매일 0시에 쿠폰을 발급하는 서비스

사용법:
    ./coupang_coupon_issuer verify [디렉토리]
    ./coupang_coupon_issuer issue [디렉토리] [--jitter-max 60]
    ./coupang_coupon_issuer install [디렉토리] --access-key KEY --secret-key SECRET ...
    ./coupang_coupon_issuer uninstall [디렉토리]
"""

import sys
import argparse
from datetime import datetime
from pathlib import Path

# 개발 시 모듈 경로 추가
sys.path.insert(0, str(Path(__file__).parent))

from coupang_coupon_issuer.config import ConfigManager, get_excel_file
from coupang_coupon_issuer.issuer import CouponIssuer
from coupang_coupon_issuer.service import CrontabService


def cmd_verify(args) -> None:
    """엑셀 검증 및 전체 내용 출력 (테이블 형식)"""
    from openpyxl import load_workbook

    base_dir = Path(args.directory).resolve() if args.directory else Path.cwd()
    excel_path = get_excel_file(base_dir)  # 항상 coupons.xlsx 사용

    if not excel_path.exists():
        print(f"ERROR: {excel_path} 파일을 찾을 수 없습니다", flush=True)
        sys.exit(1)

    print(f"엑셀 파일 검증 중: {excel_path}", flush=True)

    # 엑셀 검증 (CouponIssuer 없이 직접 읽기)
    try:
        workbook = load_workbook(excel_path, read_only=True)
        sheet = workbook.active

        if sheet is None:
            raise ValueError("엑셀 시트를 찾을 수 없습니다")

        # 헤더 읽기
        headers_row = [cell.value for cell in sheet[1]]

        # 필수 컬럼 체크
        required_columns = ['쿠폰이름', '쿠폰타입', '쿠폰유효기간', '할인방식', '할인금액/비율', '발급개수']
        for col in required_columns:
            if col not in headers_row:
                raise ValueError(f"필수 컬럼이 없습니다: {col}")

        # 데이터 행 읽기
        coupons = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 빈 행 건너뛰기
                continue

            col_indices = {header: idx for idx, header in enumerate(headers_row)}

            coupon_name = str(row[col_indices['쿠폰이름']]).strip()
            coupon_type = str(row[col_indices['쿠폰타입']]).strip()
            validity_days_raw = row[col_indices['쿠폰유효기간']]
            discount_type = str(row[col_indices['할인방식']]).strip().upper()
            discount_raw = row[col_indices['할인금액/비율']]
            issue_count_raw = row[col_indices['발급개수']]

            # 타입 변환 (openpyxl cell value 처리)
            try:
                validity_days = int(float(str(validity_days_raw))) if validity_days_raw else 0
            except (ValueError, TypeError):
                validity_days = 0

            try:
                discount = int(float(str(discount_raw))) if discount_raw else 0
            except (ValueError, TypeError):
                discount = 0

            # 발급개수 처리
            if coupon_type == '즉시할인':
                issue_count = 0
            else:
                try:
                    issue_count = int(float(str(issue_count_raw))) if issue_count_raw else 1
                except (ValueError, TypeError):
                    issue_count = 1

            coupons.append({
                'name': coupon_name,
                'type': coupon_type,
                'validity_days': validity_days,
                'discount_type': discount_type,
                'discount': discount,
                'issue_count': issue_count
            })

        workbook.close()

    except Exception as e:
        print(f"ERROR: 엑셀 로드 실패: {e}", flush=True)
        sys.exit(1)

    # 테이블 형식 출력 (엑셀처럼)
    print(f"\n✓ {len(coupons)}개 쿠폰 로드 완료\n", flush=True)

    # 헤더 (6개 컬럼 + 예산)
    headers = [
        "No", "쿠폰이름", "쿠폰타입", "유효기간", "할인방식",
        "할인금액", "할인비율", "발급개수", "총 예산"
    ]

    # 헤더 출력
    header_line = "  ".join(f"{h:^12}" for h in headers)
    print(header_line)
    print("=" * len(header_line))

    # 각 쿠폰 출력
    for i, coupon in enumerate(coupons, 1):
        # 할인금액/비율 구분
        discount_type = coupon['discount_type']
        discount = coupon['discount']

        if discount_type == 'RATE':
            discount_amount = 0
            discount_rate = discount
        else:  # PRICE, FIXED_WITH_QUANTITY
            discount_amount = discount
            discount_rate = 0

        issue_count = coupon['issue_count']

        # 예산 계산 (할인금액 × 발급개수)
        budget = discount_amount * issue_count if discount_amount > 0 else 0

        row = [
            f"{i:^12}",
            f"{coupon['name'][:10]:^12}",
            f"{coupon['type'][:10]:^12}",
            f"{coupon['validity_days']:^12}",
            f"{discount_type[:10]:^12}",
            f"{discount_amount:^12,}",
            f"{discount_rate:^12}%",
            f"{issue_count:^12,}",
            f"{budget:^12,}원"
        ]
        print("  ".join(row))

    print("\n검증 완료. 문제없이 발급 가능합니다.\n", flush=True)


def cmd_issue(args) -> None:
    """단발성 쿠폰 발급 (옵션으로 jitter 적용 가능)"""
    base_dir = Path(args.directory).resolve()

    # 1. Jitter 처리 (선택사항)
    if hasattr(args, 'jitter_max') and args.jitter_max is not None and args.jitter_max > 0:
        from coupang_coupon_issuer.jitter import JitterScheduler

        try:
            scheduler = JitterScheduler(max_jitter_minutes=args.jitter_max)
            scheduler.wait_with_jitter()
        except ValueError as e:
            print(f"ERROR: Jitter 설정 오류: {e}", flush=True)
            sys.exit(1)
        except KeyboardInterrupt:
            print("\n쿠폰 발급이 중단되었습니다.", flush=True)
            sys.exit(130)

    # 2. 쿠폰 발급 시작
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{timestamp}] 쿠폰 발급 시작...", flush=True)

    # config.json에서 API 키 로드
    try:
        ConfigManager.load_credentials_to_env(base_dir)
    except Exception as e:
        print(f"ERROR: API 키 로드 실패: {e}", flush=True)
        sys.exit(1)

    # 쿠폰 발급 실행
    try:
        issuer = CouponIssuer(base_dir)
        issuer.issue()
    except Exception as e:
        print(f"ERROR: 쿠폰 발급 실패: {e}", flush=True)
        sys.exit(1)


def cmd_install(args) -> None:
    """Cron 기반 서비스 설치"""
    base_dir = Path(args.directory).resolve()

    # # 4개 파라미터 모두 필수
    # if not all([args.access_key, args.secret_key, args.user_id, args.vendor_id]):
    #     print("ERROR: 모든 인자가 필요합니다.", flush=True)
    #     print("필수 인자:", flush=True)
    #     print("  --access-key  : Coupang Access Key", flush=True)
    #     print("  --secret-key  : Coupang Secret Key", flush=True)
    #     print("  --user-id     : WING 사용자 ID", flush=True)
    #     print("  --vendor-id   : 판매자 ID", flush=True)
    #     sys.exit(1)
    
    if not args.access_key:
        args.access_key = input("access key: ")
        
    if not args.secret_key:
        args.secret_key = input("secret key: ")
        
    if not args.user_id:
        args.user_id = input("user id: ")
        
    if not args.vendor_id:
        args.vendor_id = input("vendor id: ")

    # Jitter 범위 검증 (선택사항)
    if hasattr(args, 'jitter_max') and args.jitter_max is not None:
        if not (1 <= args.jitter_max <= 120):
            print(f"ERROR: --jitter-max는 1-120 범위여야 합니다 (현재: {args.jitter_max})", flush=True)
            sys.exit(1)

    CrontabService.install(
        base_dir,
        args.access_key,
        args.secret_key,
        args.user_id,
        args.vendor_id,
        jitter_max=args.jitter_max if hasattr(args, 'jitter_max') else None
    )


def cmd_uninstall(args) -> None:
    """Cron 기반 서비스 제거"""
    base_dir = Path(args.directory).resolve()
    CrontabService.uninstall(base_dir)


def main() -> None:
    """메인 진입점"""
    parser = argparse.ArgumentParser(
        description="Coupang Coupon Issuer - 매일 0시 쿠폰 발급 서비스",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
사용 예시:
  # 1. 엑셀 파일 검증 (coupons.xlsx 고정)
  ./coupang_coupon_issuer verify [디렉토리]

  # 2. 단발성 쿠폰 발급 (테스트용)
  ./coupang_coupon_issuer issue [디렉토리] [--jitter-max 60]

  # 3. 서비스 설치 (4개 파라미터 필수)
  ./coupang_coupon_issuer install [디렉토리] \\
    --access-key YOUR_KEY \\
    --secret-key YOUR_SECRET \\
    --user-id YOUR_USER_ID \\
    --vendor-id YOUR_VENDOR_ID \\
    [--jitter-max 60]

  # 4. 서비스 제거
  ./coupang_coupon_issuer uninstall [디렉토리]

서비스 관리:
  crontab -l                    # 스케줄 확인
  tail -f [디렉토리]/issuer.log # 로그 확인
        """
    )

    subparsers = parser.add_subparsers(dest="command", help="명령어")

    # verify 서브파서 (apply 대체)
    verify_parser = subparsers.add_parser("verify", help="엑셀 파일 검증 및 미리보기 (coupons.xlsx)")
    verify_parser.add_argument(
        "directory",
        nargs="?",
        default=None,
        help="작업 디렉토리 (기본: 현재 디렉토리)"
    )

    # issue 서브파서
    issue_parser = subparsers.add_parser("issue", help="단발성 쿠폰 발급 (즉시 실행)")
    issue_parser.add_argument(
        "directory",
        nargs="?",
        default=".",
        help="작업 디렉토리 (기본: 현재 디렉토리)"
    )
    issue_parser.add_argument(
        "--jitter-max",
        type=int,
        metavar="MINUTES",
        dest="jitter_max",
        help="최대 Jitter 시간 (분 단위, 1-120 범위)"
    )

    # install 파서
    install_parser = subparsers.add_parser("install", help="Cron 기반 서비스 설치")
    install_parser.add_argument(
        "directory",
        nargs="?",
        default=".",
        help="작업 디렉토리"
    )
    install_parser.add_argument("--access-key", help="Coupang Access Key")
    install_parser.add_argument("--secret-key", help="Coupang Secret Key")
    install_parser.add_argument("--user-id", help="WING 사용자 ID")
    install_parser.add_argument("--vendor-id", help="판매자 ID")
    install_parser.add_argument(
        "--jitter-max",
        type=int,
        metavar="MINUTES",
        help="최대 Jitter 시간 (분 단위, 1-120 범위, 기본: 미사용)"
    )

    # uninstall 파서
    uninstall_parser = subparsers.add_parser("uninstall", help="Cron 기반 서비스 제거")
    uninstall_parser.add_argument(
        "directory",
        nargs="?",
        default=".",
        help="작업 디렉토리 (기본: 현재 디렉토리)"
    )

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    # 명령어 실행
    if args.command == "verify":
        cmd_verify(args)
    elif args.command == "issue":
        cmd_issue(args)
    elif args.command == "install":
        cmd_install(args)
    elif args.command == "uninstall":
        cmd_uninstall(args)
    else:
        print(f"ERROR: 알 수 없는 명령어: {args.command}")
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
