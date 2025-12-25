#!/usr/bin/env python3
"""엑셀 예시 파일 생성 스크립트

사용법:
    python scripts/generate_example.py
    # 또는
    uv run python scripts/generate_example.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


def add_header_comments(ws):
    """헤더에 상세 설명 주석 추가 (9컬럼 구조, ADR 021)"""
    comments = {
        # Column A: 쿠폰이름
        'A1': (
            "쿠폰의 이름을 입력하세요.\n\n"
            "예시:\n"
            "- 신규회원 할인쿠폰\n"
            "- 주말 특별 할인\n"
            "- 1월 감사 쿠폰"
        ),
        
        # Column B: 쿠폰타입
        'B1': (
            "쿠폰 타입을 선택하세요.\n\n"
            "▪ 즉시할인: 상품 페이지에서 자동으로 할인 적용\n"
            "  - 옵션ID 최대 10,000개\n"
            "  - 정률할인 1~100% 가능\n\n"
            "▪ 다운로드쿠폰: 사용자가 다운로드 후 사용\n"
            "  - 옵션ID 최대 100개\n"
            "  - 정률할인 1~99%만 가능\n"
            "  - 정액할인은 10원 단위\n\n"
            "⚠ 드롭다운에서 선택하세요"
        ),
        
        # Column C: 쿠폰유효기간
        'C1': (
            "쿠폰의 유효 기간을 일(day) 단위로 입력하세요.\n\n"
            "예시:\n"
            "- 7 → 7일간 유효\n"
            "- 30 → 30일간 유효\n"
            "- 365 → 1년간 유효\n\n"
            "⚠ 1 이상의 정수만 입력"
        ),
        
        # Column D: 할인방식
        'D1': (
            "할인 방식을 선택하세요.\n\n"
            "▪ 정률할인: 정해진 비율(%)만큼 할인\n"
            "  - 다운로드쿠폰: 1~99%\n"
            "  - 즉시할인: 1~100%\n"
            "  - 최대할인금액 필수 설정\n\n"
            "▪ 정액할인: 정해진 금액(원)만큼 할인\n"
            "  - 다운로드쿠폰: 최소 10원, 10원 단위\n"
            "  - 즉시할인: 1원 이상\n\n"
            "▪ 수량별 정액할인: n개 구매 시 n번 할인\n"
            "  - 즉시할인 전용 (다운로드쿠폰 불가)\n"
            "  - 1 이상의 정수\n\n"
            "⚠ 드롭다운에서 선택하세요"
        ),
        
        # Column E: 할인금액/비율
        'E1': (
            "할인 방식에 따른 할인 값을 입력하세요.\n\n"
            "▪ 정률할인:\n"
            "  - 다운로드쿠폰: 1~99 (10% → 10 입력)\n"
            "  - 즉시할인: 1~100 (100% 가능)\n\n"
            "▪ 정액할인:\n"
            "  - 다운로드쿠폰: 10원 단위 (1000 → 1000원)\n"
            "  - 즉시할인: 1원 이상 (1원 단위 가능)\n\n"
            "▪ 수량별 정액할인:\n"
            "  - 1 이상의 정수\n\n"
            "⚠ 숫자만 입력 (%, 원 기호 제외)"
        ),
        
        # Column F: 최소구매금액 (ADR 021 - 신규)
        'F1': (
            "쿠폰 사용을 위한 최소 구매 금액을 입력하세요.\n\n"
            "▪ 다운로드쿠폰 전용:\n"
            "  - 예: 10000 → 1만원 이상 구매 시 사용 가능\n"
            "  - 빈칸 → 기본값 1원 (제한 없음)\n"
            "  - 1원 이상의 정수\n\n"
            "▪ 즉시할인쿠폰:\n"
            "  - 사용하지 않음 (비워두세요)\n\n"
            "⚠ 다운로드쿠폰만 입력, 즉시할인은 비워두기"
        ),
        
        # Column G: 최대할인금액 (ADR 021 - 신규)
        'G1': (
            "정률할인 시 최대 할인 금액을 입력하세요.\n\n"
            "▪ 모든 쿠폰 필수 입력:\n"
            "  - 예: 5000 → 최대 5,000원까지만 할인\n"
            "  - 10% 할인이어도 5,000원 초과 불가\n"
            "  - 1원 이상의 정수\n\n"
            "▪ 정액할인의 경우:\n"
            "  - 할인금액과 동일하게 설정 권장\n"
            "  - 예: 3000원 할인 → 3000 입력\n\n"
            "⚠ 필수 항목 (비워두면 오류)"
        ),
        
        # Column H: 발급개수
        'H1': (
            "다운로드쿠폰의 일일 발급 개수를 입력하세요.\n\n"
            "▪ 다운로드쿠폰:\n"
            "  - 예: 100 → 하루에 100개까지 발급\n"
            "  - 빈칸 → 기본값 1개\n"
            "  - 1 이상의 정수\n\n"
            "▪ 즉시할인쿠폰:\n"
            "  - 사용하지 않음 (비워두세요)\n\n"
            "⚠ 다운로드쿠폰만 입력, 즉시할인은 비워두기"
        ),
        
        # Column I: 옵션ID
        'I1': (
            "쿠폰을 적용할 상품 옵션 ID를 입력하세요.\n\n"
            "▪ 입력 형식:\n"
            "  - 단일: 1234567890\n"
            "  - 여러 개: 1234567890,9876543210,1122334455\n"
            "  - 쉼표로 구분, 공백 없이\n\n"
            "▪ 개수 제한 (ADR 017):\n"
            "  - 즉시할인: 최대 10,000개\n"
            "  - 다운로드쿠폰: 최대 100개\n\n"
            "⚠ 필수 항목 (양의 정수만)"
        ),
    }
    
    for cell_ref, comment_text in comments.items():
        ws[cell_ref].comment = Comment(comment_text, "시스템")


def add_data_validations(ws):
    """데이터 유효성 검사 추가 (공통 제약만, ADR 017/021 기반)"""
    
    # Column B: 쿠폰타입 (드롭다운)
    dv_coupon_type = DataValidation(
        type="list",
        formula1='"즉시할인,다운로드쿠폰"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="입력 오류",
        error="'즉시할인' 또는 '다운로드쿠폰'을 선택하세요",
        promptTitle="쿠폰타입",
        prompt="드롭다운에서 즉시할인 또는 다운로드쿠폰을 선택하세요"
    )
    dv_coupon_type.add('B2:B1000')
    ws.add_data_validation(dv_coupon_type)
    
    # Column C: 쿠폰유효기간 (1 이상 정수)
    dv_validity = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=1,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="입력 오류",
        error="1 이상의 정수를 입력하세요",
        promptTitle="쿠폰유효기간",
        prompt="유효기간을 일(day) 단위로 입력하세요 (1 이상)"
    )
    dv_validity.add('C2:C1000')
    ws.add_data_validation(dv_validity)
    
    # Column D: 할인방식 (드롭다운)
    dv_discount_type = DataValidation(
        type="list",
        formula1='"정률할인,정액할인,수량별 정액할인"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="입력 오류",
        error="'정률할인', '정액할인', '수량별 정액할인' 중 하나를 선택하세요",
        promptTitle="할인방식",
        prompt="드롭다운에서 할인 방식을 선택하세요"
    )
    dv_discount_type.add('D2:D1000')
    ws.add_data_validation(dv_discount_type)
    
    # Column E: 할인금액/비율 (1 이상 정수, 쿠폰 타입별 검증은 실행 시)
    dv_discount = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=1,
        allow_blank=False,
        showInputMessage=True,
        showErrorMessage=False,  # Warning만 (쿠폰 타입별로 검증이 다름)
        errorStyle="warning",
        errorTitle="확인 필요",
        error=(
            "쿠폰타입과 할인방식에 따라 값의 범위가 다릅니다.\n"
            "- 다운로드쿠폰 정률할인: 1~99\n"
            "- 즉시할인 정률할인: 1~100\n"
            "- 다운로드쿠폰 정액할인: 10원 단위\n"
            "- 즉시할인 정액할인: 1원 이상"
        ),
        promptTitle="할인금액/비율",
        prompt=(
            "정률할인: 1~100 (% 기호 없이)\n"
            "정액할인: 금액 (원 없이)\n"
            "수량별 정액할인: 1 이상"
        )
    )
    dv_discount.add('E2:E1000')
    ws.add_data_validation(dv_discount)
    
    # Column F: 최소구매금액 (1 이상 정수, 선택적, 다운로드쿠폰 전용)
    dv_min_purchase = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=1,
        allow_blank=True,  # 선택적 (기본값 1원)
        showInputMessage=True,
        showErrorMessage=False,  # Warning만
        errorStyle="warning",
        errorTitle="확인 필요",
        error="다운로드쿠폰인 경우 1원 이상의 정수를 입력하세요. 즉시할인은 비워두세요.",
        promptTitle="최소구매금액",
        prompt=(
            "다운로드쿠폰 전용 (즉시할인은 비워두기)\n"
            "예: 10000 → 1만원 이상 구매 시 사용 가능\n"
            "빈칸 → 기본값 1원 (제한 없음)"
        )
    )
    dv_min_purchase.add('F2:F1000')
    ws.add_data_validation(dv_min_purchase)
    
    # Column G: 최대할인금액 (1 이상 정수, 필수)
    dv_max_discount = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=1,
        allow_blank=False,  # 필수
        showInputMessage=True,
        showErrorMessage=True,
        errorTitle="입력 오류",
        error="최대할인금액은 필수입니다. 1원 이상의 정수를 입력하세요.",
        promptTitle="최대할인금액",
        prompt=(
            "정률할인 시 최대 할인 금액 (필수)\n"
            "예: 5000 → 최대 5,000원까지만 할인\n"
            "정액할인은 할인금액과 동일하게 설정 권장"
        )
    )
    dv_max_discount.add('G2:G1000')
    ws.add_data_validation(dv_max_discount)
    
    # Column H: 발급개수 (1 이상 정수, 선택적, 다운로드쿠폰 전용)
    dv_issue_count = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=1,
        allow_blank=True,  # 즉시할인은 비워둬도 됨
        showInputMessage=True,
        showErrorMessage=False,  # Warning만
        errorStyle="warning",
        errorTitle="확인 필요",
        error="다운로드쿠폰인 경우 1 이상의 숫자를 입력하세요. 즉시할인은 비워두세요.",
        promptTitle="발급개수",
        prompt=(
            "다운로드쿠폰의 일일 발급 개수 (즉시할인은 비워두기)\n"
            "예: 100 → 하루에 100개까지 발급\n"
            "빈칸 → 기본값 1개"
        )
    )
    dv_issue_count.add('H2:H1000')
    ws.add_data_validation(dv_issue_count)
    
    # Column I: 옵션ID (텍스트, 필수)
    dv_vendor_items = DataValidation(
        type="textLength",
        operator="greaterThan",
        formula1=0,
        allow_blank=False,
        showInputMessage=True,
        showErrorMessage=True,
        errorTitle="입력 오류",
        error="옵션ID는 필수 입력입니다. 숫자를 쉼표로 구분하여 입력하세요.",
        promptTitle="옵션ID",
        prompt=(
            "상품 옵션 ID를 입력하세요 (여러 개는 쉼표로 구분)\n"
            "즉시할인: 최대 10,000개\n"
            "다운로드쿠폰: 최대 100개"
        )
    )
    dv_vendor_items.add('I2:I1000')
    ws.add_data_validation(dv_vendor_items)


def create_excel_with_headers():
    """헤더가 포함된 새 워크북 생성"""
    wb = Workbook()
    ws = wb.active
    
    # 헤더 정의 (9개 컬럼)
    headers = [
        '쿠폰이름',
        '쿠폰타입',
        '쿠폰유효기간',
        '할인방식',
        '할인금액/비율',
        '최소구매금액',
        '최대할인금액',
        '발급개수',
        '옵션ID'
    ]
    
    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 컬럼 너비 조절
    ws.column_dimensions['A'].width = 20  # 쿠폰이름
    ws.column_dimensions['B'].width = 15  # 쿠폰타입
    ws.column_dimensions['C'].width = 15  # 쿠폰유효기간
    ws.column_dimensions['D'].width = 20  # 할인방식
    ws.column_dimensions['E'].width = 15  # 할인금액/비율
    ws.column_dimensions['F'].width = 15  # 최소구매금액
    ws.column_dimensions['G'].width = 15  # 최대할인금액
    ws.column_dimensions['H'].width = 12  # 발급개수
    ws.column_dimensions['I'].width = 25  # 옵션ID
    
    # 헤더 주석 추가
    add_header_comments(ws)
    
    # 데이터 유효성 검사 추가
    add_data_validations(ws)
    
    return wb, ws


def generate_basic_example(output_dir: Path):
    """기본 예제 생성 - 간단한 쿠폰 2개"""
    wb, ws = create_excel_with_headers()
    
    # 데이터 행 (9컬럼: 이름, 타입, 유효기간, 할인방식, 할인금액/비율, 최소구매금액, 최대할인금액, 발급개수, 옵션ID)
    data = [
        ['즉시할인쿠폰', '즉시할인', 30, '정률할인', 10, '', 5000, '', '1234567890,1234567891'],
        ['다운로드쿠폰', '다운로드쿠폰', 15, '정액할인', 1000, 10000, 10000, 100, '9876543210'],
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    output_path = output_dir / 'basic.xlsx'
    wb.save(output_path)
    print(f"✓ 기본 예제 생성: {output_path}")


def generate_comprehensive_example(output_dir: Path):
    """포괄적 예제 - 모든 쿠폰 타입과 할인 방식 포함"""
    wb, ws = create_excel_with_headers()
    
    # 데이터 행 (모든 조합 예시, 9컬럼)
    data = [
        # 즉시할인 쿠폰들
        ['즉시할인_정률할인', '즉시할인', 30, '정률할인', 15, '', 10000, '', '1111111111,2222222222,3333333333'],
        ['즉시할인_정액할인', '즉시할인', 60, '정액할인', 5000, '', 50000, '', '4444444444,5555555555'],
        ['즉시할인_수량할인', '즉시할인', 45, '수량별 정액할인', 2, '', 20000, '', '6666666666'],
        
        # 다운로드 쿠폰들
        ['다운로드_정률할인', '다운로드쿠폰', 7, '정률할인', 20, 10000, 5000, 50, '7777777777,8888888888'],
        ['다운로드_정액할인', '다운로드쿠폰', 14, '정액할인', 3000, 20000, 30000, 200, '9999999999'],
        ['다운로드_수량할인', '다운로드쿠폰', 30, '수량별 정액할인', 1, 5000, 10000, 150, '1010101010,2020202020'],
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    output_path = output_dir / 'comprehensive.xlsx'
    wb.save(output_path)
    print(f"✓ 포괄적 예제 생성: {output_path}")


def generate_edge_cases_example(output_dir: Path):
    """엣지 케이스 예제 - 최소/최대값 테스트"""
    wb, ws = create_excel_with_headers()
    
    # 데이터 행 (엣지 케이스, 9컬럼)
    data = [
        # 정률할인 최소/최대
        ['정률할인_최소1%', '즉시할인', 1, '정률할인', 1, '', 1000, '', '1234567890'],
        ['정률할인_최대99%', '다운로드쿠폰', 365, '정률할인', 99, 50000, 100000, 1000, '1234567891'],
        
        # 정액할인 최소
        ['정액할인_최소10원', '다운로드쿠폰', 7, '정액할인', 10, 5000, 10, 50, '1234567892'],
        ['정액할인_큰금액', '즉시할인', 90, '정액할인', 50000, '', 50000, '', '1234567893,1234567894'],
        
        # 수량할인 최소
        ['수량할인_최소1개', '다운로드쿠폰', 30, '수량별 정액할인', 1, 10000, 10000, 100, '1234567895'],
        
        # 여러 옵션ID (최대 100개까지 테스트 - 다운로드쿠폰 제한)
        ['다운로드_다중옵션', '다운로드쿠폰', 30, '정액할인', 2000, 15000, 20000, 300,
         ','.join([str(i) for i in range(1000000000, 1000000010)])],  # 10개 옵션
         
        # 즉시할인 다중 옵션 (더 많이 가능)
        ['즉시할인_다중옵션', '즉시할인', 30, '정률할인', 10, '', 10000, '',
         ','.join([str(i) for i in range(2000000000, 2000000050)])],  # 50개 옵션
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    output_path = output_dir / 'edge_cases.xlsx'
    wb.save(output_path)
    print(f"✓ 엣지 케이스 예제 생성: {output_path}")


def main():
    """메인 함수"""
    # 프로젝트 루트 디렉토리 기준으로 examples 디렉토리 생성
    project_root = Path(__file__).parent.parent
    output_dir = project_root / 'examples'
    output_dir.mkdir(exist_ok=True)
    
    print(f"\n엑셀 예시 파일 생성 중... (출력 디렉토리: {output_dir})\n")
    
    # 예제 파일들 생성
    generate_basic_example(output_dir)
    generate_comprehensive_example(output_dir)
    generate_edge_cases_example(output_dir)
    
    print(f"\n✅ 완료! 3개의 예제 파일이 생성되었습니다.")
    print(f"\n검증 명령어:")
    print(f"  uv run python main.py verify examples/")
    print()


if __name__ == '__main__':
    main()
