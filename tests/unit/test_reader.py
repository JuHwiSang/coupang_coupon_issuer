"""
Unit tests for reader.py - Excel coupon data reading and validation
"""
import pytest
from pathlib import Path
from openpyxl import Workbook

from coupang_coupon_issuer.reader import (
    fetch_coupons_from_excel,
    DISCOUNT_TYPE_KR_TO_EN,
    DISCOUNT_TYPE_EN_TO_KR
)


@pytest.mark.unit
class TestDiscountTypeConstants:
    """Test discount type mapping constants"""

    def test_kr_to_en_mapping(self):
        """Korean to English discount type mapping"""
        assert DISCOUNT_TYPE_KR_TO_EN['정률할인'] == 'RATE'
        assert DISCOUNT_TYPE_KR_TO_EN['수량별 정액할인'] == 'FIXED_WITH_QUANTITY'
        assert DISCOUNT_TYPE_KR_TO_EN['정액할인'] == 'PRICE'

    def test_en_to_kr_mapping(self):
        """English to Korean discount type mapping"""
        assert DISCOUNT_TYPE_EN_TO_KR['RATE'] == '정률할인'
        assert DISCOUNT_TYPE_EN_TO_KR['FIXED_WITH_QUANTITY'] == '수량별 정액할인'
        assert DISCOUNT_TYPE_EN_TO_KR['PRICE'] == '정액할인'

    def test_bidirectional_mapping(self):
        """Ensure bidirectional consistency"""
        for kr, en in DISCOUNT_TYPE_KR_TO_EN.items():
            assert DISCOUNT_TYPE_EN_TO_KR[en] == kr


@pytest.mark.unit
class TestFetchCouponsFromExcel:
    """Test fetch_coupons_from_excel() function"""

    def test_valid_excel_file(self, tmp_path):
        """Successfully parse valid Excel file"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰1", "즉시할인", 30, "정률할인", 10, "", 5000, "", "123456789"])
        ws.append(["테스트쿠폰2", "다운로드쿠폰", 15, "정액할인", 500, 10000, 500, 100, "987654321,111222333"])
        wb.save(excel_file)

        coupons = fetch_coupons_from_excel(excel_file)

        assert len(coupons) == 2
        
        # First coupon
        assert coupons[0]['name'] == "테스트쿠폰1"
        assert coupons[0]['type'] == "즉시할인"
        assert coupons[0]['validity_days'] == 30
        assert coupons[0]['discount_type'] == "RATE"
        assert coupons[0]['discount'] == 10
        assert coupons[0]['issue_count'] is None
        assert coupons[0]['vendor_items'] == [123456789]

        # Second coupon
        assert coupons[1]['name'] == "테스트쿠폰2"
        assert coupons[1]['type'] == "다운로드쿠폰"
        assert coupons[1]['validity_days'] == 15
        assert coupons[1]['discount_type'] == "PRICE"
        assert coupons[1]['discount'] == 500
        assert coupons[1]['issue_count'] == 100
        assert coupons[1]['vendor_items'] == [987654321, 111222333]

    def test_file_not_found(self, tmp_path):
        """Raise FileNotFoundError for non-existent file"""
        excel_file = tmp_path / "nonexistent.xlsx"
        
        with pytest.raises(FileNotFoundError) as exc_info:
            fetch_coupons_from_excel(excel_file)
        
        assert "엑셀 파일이 없습니다" in str(exc_info.value)

    def test_missing_required_columns(self, tmp_path):
        """Raise ValueError when required columns are missing"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간"])  # Missing columns
        ws.append(["쿠폰1", "즉시할인", 30])
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            fetch_coupons_from_excel(excel_file)
        
        assert "필수 컬럼이 없습니다" in str(exc_info.value)

    def test_korean_discount_types(self, tmp_path):
        """Korean discount type names are correctly mapped"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "정률할인", 10, "", 5000, "", "123"])
        ws.append(["쿠폰2", "즉시할인", 30, "정액할인", 100, "", 100, "", "456"])
        ws.append(["쿠폰3", "즉시할인", 30, "수량별 정액할인", 5, "", 5000, "", "789"])
        wb.save(excel_file)

        coupons = fetch_coupons_from_excel(excel_file)

        assert coupons[0]['discount_type'] == "RATE"
        assert coupons[1]['discount_type'] == "PRICE"
        assert coupons[2]['discount_type'] == "FIXED_WITH_QUANTITY"

    def test_invalid_discount_type(self, tmp_path):
        """Raise ValueError for invalid discount type"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "무료할인", 10, "", 5000, "", "123"])
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            fetch_coupons_from_excel(excel_file)
        
        assert "잘못된 할인방식" in str(exc_info.value)

    def test_download_coupon_rate_validation(self, tmp_path):
        """Download coupon RATE must be 1-99%"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "정률할인", 100, 10000, 20000, 100, "123"])  # 100% not allowed
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            fetch_coupons_from_excel(excel_file)
        
        assert "다운로드쿠폰 정률할인은 1~99 사이여야 합니다" in str(exc_info.value)

    def test_instant_coupon_rate_validation(self, tmp_path):
        """Instant coupon RATE can be 1-100%"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "정률할인", 100, "", 50000, "", "123"])  # 100% allowed
        wb.save(excel_file)

        coupons = fetch_coupons_from_excel(excel_file)
        assert len(coupons) == 1
        assert coupons[0]['discount'] == 100

    def test_download_coupon_price_validation(self, tmp_path):
        """Download coupon PRICE must be >=10 and in 10-won units"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "정액할인", 5, 10000, 5, 100, "123"])  # < 10
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            fetch_coupons_from_excel(excel_file)
        
        assert "다운로드쿠폰 정액할인은 최소 10원 이상이어야 합니다" in str(exc_info.value)

    def test_vendor_items_parsing(self, tmp_path):
        """Vendor items are correctly parsed from comma-separated list"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "정률할인", 10, "", 5000, "", "123, 456, 789"])
        wb.save(excel_file)

        coupons = fetch_coupons_from_excel(excel_file)
        assert coupons[0]['vendor_items'] == [123, 456, 789]

    def test_empty_rows_skipped(self, tmp_path):
        """Empty rows should be skipped"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "정률할인", 10, "", 5000, "", "123"])
        ws.append([None, None, None, None, None, None, None, None, None])  # Empty row
        ws.append(["쿠폰2", "다운로드쿠폰", 15, "정액할인", 100, 10000, 100, 50, "456"])
        wb.save(excel_file)

        coupons = fetch_coupons_from_excel(excel_file)
        assert len(coupons) == 2

    def test_input_normalization_coupon_type(self, tmp_path):
        """Coupon type with whitespace should be normalized"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉 시 할 인", 30, "정률할인", 10, "", 5000, "", "123"])
        wb.save(excel_file)

        coupons = fetch_coupons_from_excel(excel_file)
        assert coupons[0]['type'] == "즉시할인"

    def test_input_normalization_numbers(self, tmp_path):
        """Numbers with text should be extracted"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", "30일", "정률할인", "10%", "", 5000, "", "123"])
        wb.save(excel_file)

        coupons = fetch_coupons_from_excel(excel_file)
        assert coupons[0]['validity_days'] == 30
        assert coupons[0]['discount'] == 10

    def test_option_id_with_space_header(self, tmp_path):
        """옵션 ID (with space) should be accepted as 옵션ID"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "최소구매금액", "최대할인금액", "발급개수", "옵션 ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "정률할인", 10, "", 5000, "", "123"])
        wb.save(excel_file)

        coupons = fetch_coupons_from_excel(excel_file)
        assert len(coupons) == 1
        assert coupons[0]['vendor_items'] == [123]
