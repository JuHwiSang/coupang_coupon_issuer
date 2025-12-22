"""
Unit tests for main.py - CLI commands
"""
import pytest
import sys
from pathlib import Path
from unittest.mock import patch, MagicMock, call
from openpyxl import Workbook

# Add project root to sys.path to import main module
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import main


@pytest.mark.unit
class TestVerifyCommand:
    """Test 'verify' command (replaces apply)"""

    def test_verify_valid_excel(self, tmp_path, capsys):
        """Valid Excel file should pass validation and display table"""
        excel_file = tmp_path / "valid.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수"])
        ws.append(["테스트쿠폰1", "즉시할인", 30, "RATE", 10, ""])
        ws.append(["테스트쿠폰2", "다운로드쿠폰", 15, "PRICE", 500, 100])
        wb.save(excel_file)

        # Create args object
        args = MagicMock()
        args.file = str(excel_file)

        # Run command
        main.cmd_verify(args)

        # Verify output
        captured = capsys.readouterr()
        assert "엑셀 파일 검증 중" in captured.out
        assert "2개 쿠폰 로드 완료" in captured.out
        assert "테스트쿠폰1" in captured.out
        assert "테스트쿠폰2" in captured.out
        assert "검증 완료" in captured.out

    def test_verify_uses_default_path_when_no_file_arg(self, tmp_path, capsys, monkeypatch):
        """When no file argument, should use get_excel_file(cwd)"""
        # Create excel in current working directory
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수"])
        ws.append(["쿠폰1", "즉시할인", 30, "RATE", 10, ""])
        wb.save(excel_file)

        # Change to tmp_path directory
        import os
        old_cwd = os.getcwd()
        try:
            os.chdir(tmp_path)

            args = MagicMock()
            args.file = None  # No file specified
            args.directory = None  # Use current directory

            main.cmd_verify(args)

            captured = capsys.readouterr()
            assert "1개 쿠폰 로드 완료" in captured.out
        finally:
            os.chdir(old_cwd)

    def test_verify_file_not_found(self, tmp_path, capsys):
        """Non-existent file should error"""
        args = MagicMock()
        args.file = str(tmp_path / "nonexistent.xlsx")

        with pytest.raises(SystemExit):
            main.cmd_verify(args)

        captured = capsys.readouterr()
        assert "ERROR" in captured.out
        assert "찾을 수 없습니다" in captured.out

    def test_verify_missing_columns(self, tmp_path, capsys):
        """Excel missing required columns should error"""
        excel_file = tmp_path / "missing_cols.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입"])  # Missing 4 columns
        ws.append(["쿠폰1", "즉시할인"])
        wb.save(excel_file)

        args = MagicMock()
        args.file = str(excel_file)

        with pytest.raises(SystemExit):
            main.cmd_verify(args)

        captured = capsys.readouterr()
        assert "ERROR" in captured.out
        assert "필수 컬럼이 없습니다" in captured.out

    def test_verify_displays_rate_discount(self, tmp_path, capsys):
        """RATE discount should show 0 amount and X% rate"""
        excel_file = tmp_path / "rate.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수"])
        ws.append(["할인쿠폰", "즉시할인", 30, "RATE", 15, ""])
        wb.save(excel_file)

        args = MagicMock()
        args.file = str(excel_file)

        main.cmd_verify(args)

        captured = capsys.readouterr()
        assert "15" in captured.out and "%" in captured.out  # Should show rate
        assert "할인쿠폰" in captured.out

    def test_verify_displays_price_discount(self, tmp_path, capsys):
        """PRICE discount should show amount and 0% rate"""
        excel_file = tmp_path / "price.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수"])
        ws.append(["금액쿠폰", "다운로드쿠폰", 15, "PRICE", 1000, 50])
        wb.save(excel_file)

        args = MagicMock()
        args.file = str(excel_file)

        main.cmd_verify(args)

        captured = capsys.readouterr()
        assert "1,000" in captured.out  # Should show amount with comma
        assert "50,000" in captured.out  # Budget (1000 × 50)
        assert "금액쿠폰" in captured.out

    def test_verify_calculates_budget_correctly(self, tmp_path, capsys):
        """Budget should be discount_amount × issue_count"""
        excel_file = tmp_path / "budget.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수"])
        ws.append(["예산쿠폰", "다운로드쿠폰", 30, "PRICE", 500, 100])
        wb.save(excel_file)

        args = MagicMock()
        args.file = str(excel_file)

        main.cmd_verify(args)

        captured = capsys.readouterr()
        assert "50,000" in captured.out  # 500 × 100 = 50,000


@pytest.mark.unit
class TestIssueCommand:
    """Test 'issue' command"""

    def test_issue_loads_credentials_from_file(self, tmp_path, mocker):
        """Issue command should load credentials from file"""
        mock_load_creds = mocker.patch(
            'main.ConfigManager.load_credentials',
            return_value=("access", "secret", "user", "vendor")
        )
        mock_issuer_class = mocker.patch('main.CouponIssuer')
        mock_issuer = MagicMock()
        mock_issuer_class.return_value = mock_issuer

        args = MagicMock()
        args.jitter_max = None
        args.directory = str(tmp_path)

        main.cmd_issue(args)

        # Verify credentials were loaded
        mock_load_creds.assert_called_once()

        # Verify issuer was instantiated with credentials
        mock_issuer_class.assert_called_once_with(
            base_dir=tmp_path,
            access_key="access",
            secret_key="secret",
            user_id="user",
            vendor_id="vendor"
        )
        mock_issuer.issue.assert_called_once()

    def test_issue_handles_credential_error(self, tmp_path, mocker, capsys):
        """Issue should exit if credentials can't be loaded"""
        mocker.patch('main.ConfigManager.load_credentials', side_effect=FileNotFoundError("No file"))

        args = MagicMock()
        args.jitter_max = None
        args.directory = str(tmp_path)

        with pytest.raises(SystemExit):
            main.cmd_issue(args)

        captured = capsys.readouterr()
        assert "ERROR: API 키 로드 실패" in captured.out

    def test_issue_handles_issuer_error(self, tmp_path, mocker, capsys):
        """Issue should exit if issuer.issue() fails"""
        mocker.patch(
            'main.ConfigManager.load_credentials',
            return_value=("access", "secret", "user", "vendor")
        )
        mock_issuer_class = mocker.patch('main.CouponIssuer')
        mock_issuer = MagicMock()
        mock_issuer.issue.side_effect = Exception("Issuer failed")
        mock_issuer_class.return_value = mock_issuer

        args = MagicMock()
        args.jitter_max = None
        args.directory = str(tmp_path)

        with pytest.raises(SystemExit):
            main.cmd_issue(args)

        captured = capsys.readouterr()
        assert "ERROR: 쿠폰 발급 실패" in captured.out

    def test_issue_with_jitter(self, tmp_path, mocker):
        """Issue command should handle jitter parameter"""
        mocker.patch(
            'main.ConfigManager.load_credentials',
            return_value=("access", "secret", "user", "vendor")
        )
        mock_issuer_class = mocker.patch('main.CouponIssuer')
        mock_issuer = MagicMock()
        mock_issuer_class.return_value = mock_issuer

        # JitterScheduler is imported inside cmd_issue, so patch the module path
        mock_jitter_class = mocker.patch('coupang_coupon_issuer.jitter.JitterScheduler')
        mock_jitter = MagicMock()
        mock_jitter_class.return_value = mock_jitter

        args = MagicMock()
        args.jitter_max = 60
        args.directory = str(tmp_path)

        main.cmd_issue(args)

        # Verify jitter was used
        mock_jitter_class.assert_called_once_with(max_jitter_minutes=60)
        mock_jitter.wait_with_jitter.assert_called_once()

        # Verify issue was still called
        mock_issuer.issue.assert_called_once()


@pytest.mark.unit
class TestInstallCommand:
    """Test 'install' command"""

    def test_install_requires_all_4_params(self, tmp_path, capsys):
        """Install should require all 4 parameters"""
        args = MagicMock()
        args.access_key = "test"
        args.secret_key = "test"
        args.user_id = None  # Missing
        args.vendor_id = "test"
        args.directory = str(tmp_path)

        with pytest.raises(SystemExit):
            main.cmd_install(args)

        captured = capsys.readouterr()
        assert "ERROR: 모든 인자가 필요합니다" in captured.out
        assert "--user-id" in captured.out

    def test_install_calls_crontab_service(self, tmp_path, mocker):
        """Install should call CrontabService.install with correct args"""
        mock_install = mocker.patch('main.CrontabService.install')

        args = MagicMock()
        args.access_key = "access-key"
        args.secret_key = "secret-key"
        args.user_id = "user-id"
        args.vendor_id = "vendor-id"
        args.jitter_max = None
        args.directory = str(tmp_path)

        main.cmd_install(args)

        # CrontabService.install now takes base_dir as first argument
        from pathlib import Path
        mock_install.assert_called_once_with(Path(str(tmp_path)), "access-key", "secret-key", "user-id", "vendor-id", jitter_max=None)

    def test_install_with_jitter(self, tmp_path, mocker):
        """Install should pass jitter_max to CrontabService"""
        mock_install = mocker.patch('main.CrontabService.install')

        args = MagicMock()
        args.access_key = "access-key"
        args.secret_key = "secret-key"
        args.user_id = "user-id"
        args.vendor_id = "vendor-id"
        args.jitter_max = 60
        args.directory = str(tmp_path)

        main.cmd_install(args)

        from pathlib import Path
        mock_install.assert_called_once_with(Path(str(tmp_path)), "access-key", "secret-key", "user-id", "vendor-id", jitter_max=60)

    def test_install_validates_jitter_range(self, tmp_path, capsys):
        """Install should validate jitter_max is in 1-120 range"""
        args = MagicMock()
        args.access_key = "test"
        args.secret_key = "test"
        args.user_id = "test"
        args.vendor_id = "test"
        args.jitter_max = 150  # Out of range
        args.directory = str(tmp_path)

        with pytest.raises(SystemExit):
            main.cmd_install(args)

        captured = capsys.readouterr()
        assert "ERROR" in captured.out
        assert "1-120 범위" in captured.out


@pytest.mark.unit
class TestUninstallCommand:
    """Test 'uninstall' command"""

    def test_uninstall_calls_crontab_service(self, tmp_path, mocker):
        """Uninstall should call CrontabService.uninstall"""
        mock_uninstall = mocker.patch('main.CrontabService.uninstall')

        args = MagicMock()
        args.directory = str(tmp_path)

        main.cmd_uninstall(args)

        from pathlib import Path
        mock_uninstall.assert_called_once_with(Path(str(tmp_path)))


@pytest.mark.unit
class TestMainFunction:
    """Test main() entry point and argument parsing"""

    def test_main_no_arguments_shows_help(self, mocker, capsys):
        """Running without arguments should show help and exit"""
        mocker.patch('sys.argv', ['main.py'])

        with pytest.raises(SystemExit):
            main.main()

        captured = capsys.readouterr()
        assert "usage:" in captured.out or "사용 예시:" in captured.out

    def test_main_verify_command_dispatched(self, mocker, tmp_path):
        """Verify command should be dispatched correctly"""
        mock_cmd_verify = mocker.patch('main.cmd_verify')

        # Create valid Excel file
        excel_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수"])
        ws.append(["쿠폰", "즉시할인", 30, "RATE", 10, ""])
        wb.save(excel_file)

        mocker.patch('sys.argv', ['main.py', 'verify', str(excel_file)])
        main.main()

        # Verify cmd_verify was called
        assert mock_cmd_verify.call_count == 1

    def test_main_issue_command_dispatched(self, mocker):
        """Issue command should be dispatched correctly"""
        mock_cmd_issue = mocker.patch('main.cmd_issue')

        mocker.patch('sys.argv', ['main.py', 'issue'])
        main.main()

        mock_cmd_issue.assert_called_once()

    def test_main_install_command_dispatched(self, tmp_path, mocker):
        """Install command should be dispatched correctly"""
        mock_cmd_install = mocker.patch('main.cmd_install')

        mocker.patch('sys.argv', [
            'main.py', 'install', str(tmp_path),
            '--access-key', 'test-access',
            '--secret-key', 'test-secret',
            '--user-id', 'test-user',
            '--vendor-id', 'test-vendor'
        ])
        main.main()

        assert mock_cmd_install.call_count == 1

    def test_main_uninstall_command_dispatched(self, mocker):
        """Uninstall command should be dispatched correctly"""
        mock_cmd_uninstall = mocker.patch('main.cmd_uninstall')

        mocker.patch('sys.argv', ['main.py', 'uninstall'])
        main.main()

        mock_cmd_uninstall.assert_called_once()
