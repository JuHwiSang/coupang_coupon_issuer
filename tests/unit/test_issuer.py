"""
Unit tests for issuer.py - CouponIssuer
"""
import pytest
import os
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch, MagicMock, mock_open
from openpyxl import Workbook

from coupang_coupon_issuer.issuer import CouponIssuer
from coupang_coupon_issuer.config import (
    COUPON_MAX_DISCOUNT,
    COUPON_CONTRACT_ID,
    COUPON_DEFAULT_ISSUE_COUNT
)


@pytest.mark.unit
class TestCouponIssuerInit:
    """Test CouponIssuer initialization"""

    def test_init_with_explicit_credentials(self, tmp_path):
        """Initialize with explicit credential parameters"""
        issuer = CouponIssuer(
            base_dir=tmp_path,
            access_key="test-access",
            secret_key="test-secret",
            user_id="test-user",
            vendor_id="test-vendor"
        )

        assert issuer.access_key == "test-access"
        assert issuer.secret_key == "test-secret"
        assert issuer.user_id == "test-user"
        assert issuer.vendor_id == "test-vendor"
        assert issuer.base_dir == tmp_path



    def test_init_missing_api_keys(self, tmp_path):
        """Raise ValueError when API keys missing"""
        with pytest.raises(ValueError) as exc_info:
            CouponIssuer(
                base_dir=tmp_path,
                access_key=None,
                secret_key=None,
                user_id="user",
                vendor_id="vendor"
            )

        assert "API 키가 설정되지 않았습니다" in str(exc_info.value)

    def test_init_missing_coupon_info(self, tmp_path):
        """Raise ValueError when user_id or vendor_id missing"""
        with pytest.raises(ValueError) as exc_info:
            CouponIssuer(
                base_dir=tmp_path,
                access_key="access",
                secret_key="secret",
                user_id=None,
                vendor_id=None
            )

        assert "쿠폰 정보가 설정되지 않았습니다" in str(exc_info.value)


@pytest.mark.unit
class TestExcelParsing:
    """Test _fetch_coupons_from_excel() method"""

    def test_fetch_coupons_valid_file(self, tmp_path):
        """Read valid Excel file with 6 columns"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰1", "즉시할인", 30, "RATE", 10, "", "123456789"])  # 즉시할인: 발급개수 빈값
        ws.append(["테스트쿠폰2", "다운로드쿠폰", 15, "PRICE", 500, 100, "987654321, 111222333"])  # 다운로드: 할인 + 발급개수
        wb.save(excel_file)

        issuer = CouponIssuer(
            base_dir=tmp_path,
            access_key="test-access",
            secret_key="test-secret",
            user_id="test-user",
            vendor_id="test-vendor"
        )

        coupons = issuer._fetch_coupons_from_excel()

        assert len(coupons) == 2

        assert coupons[0]['name'] == "테스트쿠폰1"
        assert coupons[0]['type'] == "즉시할인"
        assert coupons[0]['validity_days'] == 30
        assert coupons[0]['discount_type'] == "RATE"
        assert coupons[0]['discount'] == 10  # Column E
        assert coupons[0]['issue_count'] is None  # 즉시할인: None
        assert coupons[0]['vendor_items'] == [123456789]  # Column G

        assert coupons[1]['name'] == "테스트쿠폰2"
        assert coupons[1]['type'] == "다운로드쿠폰"
        assert coupons[1]['validity_days'] == 15
        assert coupons[1]['discount_type'] == "PRICE"
        assert coupons[1]['discount'] == 500  # Column E
        assert coupons[1]['issue_count'] == 100  # Column F
        assert coupons[1]['vendor_items'] == [987654321, 111222333]  # Column G

    def test_fetch_coupons_missing_columns(self, tmp_path):
        """Raise ValueError when required columns missing"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간"])  # Missing 2 columns
        ws.append(["쿠폰1", "즉시할인", 30])
        wb.save(excel_file)

        issuer = CouponIssuer(
            base_dir=tmp_path,
            access_key="test-access",
            secret_key="test-secret",
            user_id="test-user",
            vendor_id="test-vendor"
        )

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()

        assert "필수 컬럼이 없습니다" in str(exc_info.value)

    def test_fetch_coupons_empty_rows_skipped(self, tmp_path):
        """Verify empty rows are skipped"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "RATE", 50, "", "123456789"])
        ws.append([None, None, None, None, None, None, None])  # Empty row
        ws.append(["쿠폰2", "다운로드쿠폰", 15, "PRICE", 1000, 100, "987654321"])
        wb.save(excel_file)

        issuer = CouponIssuer(
            base_dir=tmp_path,
            access_key="test-access",
            secret_key="test-secret",
            user_id="test-user",
            vendor_id="test-vendor"
        )

        coupons = issuer._fetch_coupons_from_excel()

        # Only 2 coupons, empty row skipped
        assert len(coupons) == 2

    def test_fetch_coupons_file_not_found(self, tmp_path):
        """Raise FileNotFoundError when Excel file doesn't exist"""
        nonexistent_file = tmp_path / "nonexistent.xlsx"

        issuer = CouponIssuer(
            base_dir=tmp_path,
            access_key="test-access",
            secret_key="test-secret",
            user_id="test-user",
            vendor_id="test-vendor"
        )

        with pytest.raises(FileNotFoundError) as exc_info:
            issuer._fetch_coupons_from_excel()

        assert "엑셀 파일이 없습니다" in str(exc_info.value)


@pytest.mark.unit
class TestInputNormalization:
    """Test input normalization logic (ADR 002)"""

    def test_normalize_coupon_name_strip(self, tmp_path):
        """Coupon name should be stripped of whitespace"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["  테스트쿠폰  ", "즉시할인", 30, "RATE", 50, "", "123456789"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()

        assert coupons[0]['name'] == "테스트쿠폰"  # Stripped

    def test_normalize_coupon_type_whitespace_removal(self, tmp_path):
        """Coupon type whitespace should be removed"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉 시 할 인", 30, "RATE", 50, "", "123456789"])
        ws.append(["쿠폰2", "다운로 드쿠폰", 15, "PRICE", 1000, 100, "987654321"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()

        assert coupons[0]['type'] == "즉시할인"
        assert coupons[1]['type'] == "다운로드쿠폰"

    def test_normalize_validity_period_extract_numbers(self, tmp_path):
        """Extract numbers from validity period strings"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", "30일", "RATE", 50, "", "123456789"])
        ws.append(["쿠폰2", "다운로드쿠폰", "15 days", "PRICE", 1000, 100, "987654321"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()

        assert coupons[0]['validity_days'] == 30
        assert coupons[1]['validity_days'] == 15

    def test_normalize_discount_method_uppercase(self, tmp_path):
        """Discount method should be normalized to uppercase"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "rate", 50, "", "123456789"])
        ws.append(["쿠폰2", "다운로드쿠폰", 15, "fixed-with-quantity", 3, 1000, "987654321"])
        ws.append(["쿠폰3", "즉시할인", 7, "price", 100, "", "111222333"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()

        assert coupons[0]['discount_type'] == "RATE"
        assert coupons[1]['discount_type'] == "FIXED_WITH_QUANTITY"
        assert coupons[2]['discount_type'] == "PRICE"

    def test_normalize_quantity_extract_numbers(self, tmp_path):
        """Extract numbers from discount and issue count strings"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "RATE", "10%", "", "123456789"])
        ws.append(["쿠폰2", "다운로드쿠폰", 15, "PRICE", "500 원", "100개", "987654321"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()

        # Column E: 할인금액/비율
        assert coupons[0]['discount'] == 10  # "10%" → 10
        assert coupons[1]['discount'] == 500  # "500 원" → 500

        # Column F: 발급개수
        assert coupons[0]['issue_count'] is None  # 즉시할인: None
        assert coupons[1]['issue_count'] == 100  # "100개" → 100


@pytest.mark.unit
class TestValidation:
    """Test validation rules"""

    def test_validate_download_coupon_rate_range_1_to_99(self, tmp_path):
        """Download coupon RATE must be between 1-99 (100 not allowed)"""
        # Valid cases
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "RATE", 1, 100, "123456789"])
        ws.append(["쿠폰2", "다운로드쿠폰", 30, "RATE", 50, 100, "123456789"])
        ws.append(["쿠폰3", "다운로드쿠폰", 30, "RATE", 99, 100, "123456789"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()
        assert len(coupons) == 3  # All valid

        # Invalid: 0%
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "RATE", 0, 100, "123456789"])
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()
        # 0 value triggers the general "> 0" validation before RATE-specific validation
        assert "할인금액/비율은 0보다 커야 합니다" in str(exc_info.value)

        # Invalid: 100% (download coupon doesn't allow 100)
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "RATE", 100, 100, "123456789"])
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()
        assert "다운로드쿠폰 RATE 할인율은 1~99 사이여야 합니다" in str(exc_info.value)

    def test_validate_instant_coupon_rate_range_1_to_100(self, tmp_path):
        """Instant coupon RATE can be 1-100 (100 allowed)"""
        # Valid cases including 100%
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "RATE", 1, "", "123456789"])
        ws.append(["쿠폰2", "즉시할인", 30, "RATE", 50, "", "123456789"])
        ws.append(["쿠폰3", "즉시할인", 30, "RATE", 99, "", "123456789"])
        ws.append(["쿠폰4", "즉시할인", 30, "RATE", 100, "", "123456789"])  # 100% allowed for instant
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()
        assert len(coupons) == 4  # All valid including 100%

        # Invalid: 101%
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "RATE", 101, "", "123456789"])
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()
        assert "즉시할인쿠폰 RATE 할인율은 1~100 사이여야 합니다" in str(exc_info.value)

    def test_validate_download_coupon_price_minimum_10_won(self, tmp_path):
        """Download coupon PRICE must be >= 10 won and in 10-won units"""
        # Valid
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "PRICE", 10, 100, "123456789"])
        ws.append(["쿠폰2", "다운로드쿠폰", 30, "PRICE", 100, 100, "123456789"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()
        assert len(coupons) == 2

        # Invalid: < 10
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "PRICE", 5, 100, "123456789"])
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()
        assert "다운로드쿠폰 PRICE 할인금액은 최소 10원 이상이어야 합니다" in str(exc_info.value)

    def test_validate_download_coupon_price_10_won_units(self, tmp_path):
        """Download coupon PRICE must be in 10-won units"""
        # Valid
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "PRICE", 10, 100, "123456789"])
        ws.append(["쿠폰2", "다운로드쿠폰", 30, "PRICE", 20, 100, "123456789"])
        ws.append(["쿠폰3", "다운로드쿠폰", 30, "PRICE", 100, 100, "123456789"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()
        assert len(coupons) == 3

        # Invalid: not 10-won unit
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "다운로드쿠폰", 30, "PRICE", 15, 100, "123456789"])
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()
        assert "다운로드쿠폰 PRICE 할인금액은 10원 단위여야 합니다" in str(exc_info.value)

    def test_validate_instant_coupon_price_minimum_1_won(self, tmp_path):
        """Instant coupon PRICE must be >= 1 won (no 10-won unit requirement)"""
        # Valid cases including non-10-won units
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "PRICE", 1, "", "123456789"])
        ws.append(["쿠폰2", "즉시할인", 30, "PRICE", 5, "", "123456789"])  # 5 won allowed
        ws.append(["쿠폰3", "즉시할인", 30, "PRICE", 15, "", "123456789"])  # 15 won allowed
        ws.append(["쿠폰4", "즉시할인", 30, "PRICE", 123, "", "123456789"])  # 123 won allowed
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupons = issuer._fetch_coupons_from_excel()
        assert len(coupons) == 4  # All valid

        # Invalid: 0 won
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "즉시할인", 30, "PRICE", 0, "", "123456789"])
        wb.save(excel_file)

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()
        # 0 triggers general validation first
        assert "할인금액/비율은 0보다 커야 합니다" in str(exc_info.value)


@pytest.mark.unit
class TestIssuanceWorkflow:
    """Test issue() and _issue_single_coupon() methods"""

    def test_issue_single_coupon_instant_discount(self, tmp_path, requests_mock):
        """Test instant discount coupon creation"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["즉시할인쿠폰", "즉시할인", 30, "PRICE", 1000, "", "123456789"])
        wb.save(excel_file)

        # Mock API
        requests_mock.post(
            "https://api-gateway.coupang.com/v2/providers/fms/apis/api/v2/vendors/test-vendor/coupon",
            status_code=200,
            json={"code": 200, "data": {"requestedId": "instant-123"}}
        )

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "test-vendor")

        issuer.issue()

        # Verify API was called
        assert requests_mock.call_count == 1

        # Verify request payload
        request_body = requests_mock.last_request.json()
        assert request_body["name"] == "즉시할인쿠폰"
        assert request_body["contractId"] == "-1"
        assert request_body["maxDiscountPrice"] == str(COUPON_MAX_DISCOUNT)
        assert request_body["discount"] == "1000"
        assert request_body["type"] == "PRICE"

    def test_issue_single_coupon_download_coupon(self, tmp_path, requests_mock):
        """Test download coupon creation"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["다운로드쿠폰", "다운로드쿠폰", 15, "RATE", 50, 100, "123456789"])
        wb.save(excel_file)

        # Mock API
        requests_mock.post(
            "https://api-gateway.coupang.com/v2/providers/marketplace_openapi/apis/api/v1/coupons",
            status_code=200,
            json={"code": 200, "data": {"couponId": "download-456"}}
        )

        issuer = CouponIssuer(tmp_path, "a", "s", "test-user", "v")

        issuer.issue()

        # Verify API was called
        assert requests_mock.call_count == 1

        # Verify request payload
        request_body = requests_mock.last_request.json()
        assert request_body["title"] == "다운로드쿠폰"
        assert request_body["contractId"] == COUPON_CONTRACT_ID
        assert request_body["couponType"] == "DOWNLOAD"
        assert request_body["userId"] == "test-user"
        assert len(request_body["policies"]) == 1
        assert request_body["policies"][0]["discount"] == 50

    def test_issue_handles_api_failure(self, tmp_path, requests_mock, capsys):
        """Test error handling when API call fails"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["실패쿠폰", "즉시할인", 30, "PRICE", 1000, "", "123456789"])
        wb.save(excel_file)

        # Mock API failure
        requests_mock.post(
            "https://api-gateway.coupang.com/v2/providers/fms/apis/api/v2/vendors/test-vendor/coupon",
            status_code=200,
            json={"code": 500, "errorMessage": "Internal error"}
        )

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "test-vendor")

        issuer.issue()

        # Verify log contains [FAIL] marker
        captured = capsys.readouterr()
        assert "[FAIL]" in captured.out
        assert "실패쿠폰" in captured.out

    def test_issue_mixed_success_failure(self, tmp_path, capsys):
        """Test mixed success and failure scenario"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["성공쿠폰1", "즉시할인", 30, "PRICE", 1000, "", "123456789"])
        ws.append(["성공쿠폰2", "즉시할인", 30, "PRICE", 2000, "", "123456789"])
        ws.append(["성공쿠폰3", "즉시할인", 30, "PRICE", 3000, "", "123456789"])
        ws.append(["실패쿠폰", "즉시할인", 30, "PRICE", 4000, "", "123456789"])
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "test-vendor")

        # Mock _issue_single_coupon to simulate mixed success/failure
        def mock_issue(idx, coupon):
            if coupon['name'] == "실패쿠폰":
                return {
                    'coupon_name': coupon['name'],
                    'coupon_type': coupon['type'],
                    'status': '실패',
                    'message': 'Simulated failure',
                    'timestamp': issuer._timestamp()
                }
            else:
                return {
                    'coupon_name': coupon['name'],
                    'coupon_type': coupon['type'],
                    'status': '성공',
                    'message': 'Simulated success',
                    'timestamp': issuer._timestamp()
                }

        with patch.object(issuer, '_issue_single_coupon', side_effect=mock_issue):
            issuer.issue()

        # Verify summary log
        captured = capsys.readouterr()
        assert "성공: 3, 실패: 1" in captured.out
        assert "[OK] 성공쿠폰1" in captured.out
        assert "[OK] 성공쿠폰2" in captured.out
        assert "[OK] 성공쿠폰3" in captured.out
        assert "[FAIL] 실패쿠폰" in captured.out



@pytest.mark.unit
class TestIssuerEdgeCases:
    """
    Test edge cases and missing coverage areas in issuer.py.

    Covers missing lines:
    - Lines 76-77: Empty coupon list handling
    - Lines 97-99: Exception propagation in issue()
    - Line 130: Invalid discount validation
    - Line 192: Unknown coupon type handling
    - Line 235: Empty sheet handling
    - Lines 269-281: Invalid coupon type validation
    - Lines 294-304: Invalid discount method validation
    - Lines 324-328: Invalid issue count for downloadable coupons
    - Lines 343-344: FIXED_WITH_QUANTITY validation
    """

    @pytest.fixture
    def empty_excel(self, tmp_path):
        """Excel with headers only, no data rows - returns tmp_path"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        # No data rows
        wb.save(excel_file)
        return tmp_path  # Return tmp_path so CouponIssuer can use it as base_dir

    @pytest.fixture
    def malformed_excel(self, tmp_path):
        """Excel with intentionally bad data"""
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["쿠폰1", "잘못된타입", 30, "INVALID", -10, "abc", "123456789"])
        wb.save(excel_file)
        return excel_file

    @pytest.fixture
    def coupon_dict_factory(self):
        """Factory to create coupon dicts for testing"""
        def _create(**overrides):
            default = {
                'name': '테스트쿠폰',
                'type': '즉시할인',
                'validity_days': 30,
                'discount_type': 'RATE',
                'discount': 10,
                'issue_count': None,
                'vendor_items': [123456789]
            }
            default.update(overrides)
            return default

        return _create

    def test_issue_with_empty_coupon_list(self, empty_excel, capsys):
        """
        Test issue() with empty coupon list (헤더만 있고 데이터 없음).

        Covers: issuer.py lines 76-77
        Expected: "발급할 쿠폰이 없습니다" 출력, 조기 리턴
        """
        issuer = CouponIssuer(empty_excel, "a", "s", "u", "v")

        issuer.issue()

        captured = capsys.readouterr()
        assert "발급할 쿠폰이 없습니다" in captured.out

    def test_issue_propagates_excel_read_exception(self, tmp_path):
        """
        Test issue() propagates exception from _fetch_coupons_from_excel().

        Covers: issuer.py lines 97-99
        Expected: 에러 로그 후 예외 재발생
        """
        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        # Mock _fetch_coupons_from_excel to raise exception
        with patch.object(issuer, '_fetch_coupons_from_excel', side_effect=RuntimeError("Excel error")):
            with pytest.raises(RuntimeError) as exc_info:
                issuer.issue()

            assert "Excel error" in str(exc_info.value)

    def test_issue_propagates_api_exception(self, tmp_path, requests_mock, capsys):
        """
        Test issue() handles API exception gracefully.

        Covers: issuer.py lines 196-199
        Expected: 예외를 잡아서 로그 출력, 쿠폰 발급은 계속 진행
        """
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰", "즉시할인", 30, "RATE", 10, "", "123456789"])
        wb.save(excel_file)

        # Mock API to raise connection error (correct endpoint)
        requests_mock.post(
            "https://api-gateway.coupang.com/v2/providers/fms/apis/api/v2/vendors/v/coupon",
            exc=Exception("Connection timeout")
        )

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        # Should not raise - exception is caught and logged
        issuer.issue()

        captured = capsys.readouterr()
        assert "실패: 1" in captured.out
        assert "Connection timeout" in captured.out

    def test_issue_single_coupon_raises_on_zero_discount(self, tmp_path, coupon_dict_factory):
        """
        Test _issue_single_coupon() with discount=0.

        Covers: issuer.py line 130
        Expected: ValueError 발생, 쿠폰명 포함
        """
        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupon = coupon_dict_factory(discount=0)

        with pytest.raises(ValueError) as exc_info:
            issuer._issue_single_coupon(1, coupon)

        assert "할인금액/비율이 설정되지 않았습니다" in str(exc_info.value)
        assert "테스트쿠폰" in str(exc_info.value)

    def test_issue_single_coupon_unknown_type(self, tmp_path, coupon_dict_factory):
        """
        Test _issue_single_coupon() with invalid coupon type.

        Covers: issuer.py line 192
        Expected: 실패 상태, "알 수 없는 쿠폰 타입" 메시지
        """
        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        coupon = coupon_dict_factory(type='잘못된타입')

        result = issuer._issue_single_coupon(1, coupon)

        assert result['status'] == '실패'
        assert "알 수 없는 쿠폰 타입" in result['message']
        assert "잘못된타입" in result['message']

    def test_fetch_coupons_from_empty_workbook(self, tmp_path):
        """
        Test _fetch_coupons_from_excel() with workbook that has no active sheet after load.

        Covers: issuer.py line 235
        Expected: ValueError: "시트를 찾을 수 없습니다"
        """
        excel_file = tmp_path / "coupons.xlsx"

        # Create minimal valid workbook (openpyxl requires at least one visible sheet to save)
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        # Mock load_workbook to return a workbook with no active sheet
        with patch('coupang_coupon_issuer.issuer.load_workbook') as mock_load:
            mock_wb = MagicMock()
            mock_wb.active = None
            mock_load.return_value = mock_wb

            with pytest.raises(ValueError) as exc_info:
                issuer._fetch_coupons_from_excel()

            assert "시트를 찾을 수 없습니다" in str(exc_info.value)

    def test_fetch_coupons_invalid_coupon_type(self, tmp_path):
        """
        Test _fetch_coupons_from_excel() with invalid coupon type.

        Covers: issuer.py lines 269-271
        Expected: ValueError: "잘못된 쿠폰 타입"
        """
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰", "무료쿠폰", 30, "RATE", 10, "", "123456789"])  # Invalid type
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()

        assert "잘못된 쿠폰 타입" in str(exc_info.value)
        assert "즉시할인 또는 다운로드쿠폰만 가능" in str(exc_info.value)

    def test_fetch_coupons_invalid_discount_method(self, tmp_path):
        """
        Test _fetch_coupons_from_excel() with invalid discount method.

        Covers: issuer.py lines 294-296
        Expected: ValueError: "잘못된 할인방식"
        """
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰", "즉시할인", 30, "FREE", 10, "", "123456789"])  # Invalid method
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()

        assert "잘못된 할인방식" in str(exc_info.value)
        assert "RATE/FIXED_WITH_QUANTITY/PRICE만 가능" in str(exc_info.value)

    def test_fetch_coupons_non_numeric_discount(self, tmp_path):
        """
        Test _fetch_coupons_from_excel() with non-numeric discount value.

        Covers: issuer.py lines 302-307
        Expected: ValueError with message about discount being 0 or less
        (non-numeric strings like "abc" become 0 after regex extraction)
        """
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰", "즉시할인", 30, "RATE", "abc", "", "123456789"])  # Non-numeric becomes 0
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()

        # "abc" is stripped to "" by regex, converted to 0, caught by > 0 check
        assert "할인금액/비율은 0보다 커야 합니다" in str(exc_info.value)

    def test_fetch_coupons_downloadable_with_zero_issue_count(self, tmp_path):
        """
        Test _fetch_coupons_from_excel() with downloadable coupon and issue_count=0.

        Covers: issuer.py lines 324-328
        Expected: ValueError: "1 이상이어야 합니다"
        """
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰", "다운로드쿠폰", 30, "RATE", 10, 0, "123456789"])  # issue_count=0
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()

        assert "발급개수는 1 이상이어야 합니다" in str(exc_info.value)

    def test_fetch_coupons_downloadable_with_invalid_issue_count(self, tmp_path):
        """
        Test _fetch_coupons_from_excel() with downloadable coupon and non-numeric issue_count.

        Covers: issuer.py lines 320-329 (input normalization)
        Expected: Non-numeric strings like "xyz" are normalized to empty string,
                  which triggers default value usage (no exception)
        """
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰", "다운로드쿠폰", 30, "RATE", 10, "xyz", "123456789"])  # Non-numeric -> default value
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        # Should not raise - "xyz" becomes empty string, uses default value
        coupons = issuer._fetch_coupons_from_excel()

        assert len(coupons) == 1
        assert coupons[0]['issue_count'] == COUPON_DEFAULT_ISSUE_COUNT

    def test_validate_fixed_with_quantity_minimum(self, tmp_path):
        """
        Test _fetch_coupons_from_excel() with FIXED_WITH_QUANTITY and discount=0.

        Covers: issuer.py lines 306-307 (general discount > 0 check first)
        Expected: ValueError from line 307 catches this before FIXED_WITH_QUANTITY check
        """
        excel_file = tmp_path / "coupons.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "할인금액/비율", "발급개수", "옵션ID"])
        ws.append(["테스트쿠폰", "즉시할인", 30, "FIXED_WITH_QUANTITY", 0, "", "123456789"])  # discount=0
        wb.save(excel_file)

        issuer = CouponIssuer(tmp_path, "a", "s", "u", "v")

        with pytest.raises(ValueError) as exc_info:
            issuer._fetch_coupons_from_excel()

        # General check (line 307) happens before FIXED_WITH_QUANTITY check (line 344)
        assert "할인금액/비율은 0보다 커야 합니다" in str(exc_info.value)
