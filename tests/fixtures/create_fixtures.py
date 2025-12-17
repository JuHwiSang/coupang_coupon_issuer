"""
Script to create fixture Excel files for testing
"""
from openpyxl import Workbook
from pathlib import Path

fixtures_dir = Path(__file__).parent

# 1. Valid Excel file
wb = Workbook()
ws = wb.active
assert ws is not None
ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "발급개수"])
ws.append(["할인쿠폰1", "즉시할인", 30, "RATE", 1000])
ws.append(["할인쿠폰2", "다운로드쿠폰", 15, "PRICE", 500])
ws.append(["할인쿠폰3", "즉시할인", 7, "RATE", 2000])
wb.save(fixtures_dir / "sample_valid.xlsx")
print("Created sample_valid.xlsx")

# 2. Invalid columns (missing column)
wb = Workbook()
ws = wb.active
assert ws is not None
ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식"])  # Missing "발급개수"
ws.append(["잘못된쿠폰", "즉시할인", 30, "RATE"])
wb.save(fixtures_dir / "sample_invalid_columns.xlsx")
print("Created sample_invalid_columns.xlsx")

# 3. Invalid rates (RATE out of 1-99 range)
wb = Workbook()
ws = wb.active
assert ws is not None
ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "발급개수"])
ws.append(["비율오류1", "즉시할인", 30, "RATE", 0])      # 0% - invalid
ws.append(["비율오류2", "즉시할인", 30, "RATE", 100])    # 100% - invalid
ws.append(["비율오류3", "즉시할인", 30, "RATE", -5])     # negative - invalid
wb.save(fixtures_dir / "sample_invalid_rates.xlsx")
print("Created sample_invalid_rates.xlsx")

# 4. Invalid prices (PRICE not in 10-won units)
wb = Workbook()
ws = wb.active
assert ws is not None
ws.append(["쿠폰이름", "쿠폰타입", "쿠폰유효기간", "할인방식", "발급개수"])
ws.append(["금액오류1", "다운로드쿠폰", 15, "PRICE", 5])     # < 10 won
ws.append(["금액오류2", "다운로드쿠폰", 15, "PRICE", 15])    # not 10-won unit
ws.append(["금액오류3", "다운로드쿠폰", 15, "PRICE", 105])   # not 10-won unit
wb.save(fixtures_dir / "sample_invalid_prices.xlsx")
print("Created sample_invalid_prices.xlsx")

print("\nAll fixture files created successfully!")
