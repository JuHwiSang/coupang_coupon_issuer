"""쿠폰 발급 로직 모듈"""

import os
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any

from openpyxl import load_workbook

from .coupang_api import CoupangAPIClient
from .config import (
    EXCEL_INPUT_FILE,
    COUPON_MAX_DISCOUNT,
    COUPON_CONTRACT_ID,
    COUPON_DEFAULT_ISSUE_COUNT,
)


class CouponIssuer:
    """쿠폰 발급 담당 클래스"""

    def __init__(
        self,
        access_key: Optional[str] = None,
        secret_key: Optional[str] = None,
        user_id: Optional[str] = None,
        vendor_id: Optional[str] = None,
    ):
        """
        Args:
            access_key: Coupang Access Key (None이면 환경변수에서 가져옴)
            secret_key: Coupang Secret Key (None이면 환경변수에서 가져옴)
            user_id: WING 사용자 ID (None이면 환경변수에서 가져옴)
            vendor_id: 판매자 ID (None이면 환경변수에서 가져옴)
        """
        self.access_key = access_key or os.environ.get("COUPANG_ACCESS_KEY")
        self.secret_key = secret_key or os.environ.get("COUPANG_SECRET_KEY")
        self.user_id = user_id or os.environ.get("COUPANG_USER_ID")
        self.vendor_id = vendor_id or os.environ.get("COUPANG_VENDOR_ID")

        if not self.access_key or not self.secret_key:
            raise ValueError("API 키가 설정되지 않았습니다.")

        if not self.user_id or not self.vendor_id:
            raise ValueError("쿠폰 정보가 설정되지 않았습니다 (user_id, vendor_id).")

        # Coupang API 클라이언트 초기화
        self.api_client = CoupangAPIClient(self.access_key, self.secret_key)

        print(f"[{self._timestamp()}] 설정 로드 완료 (Vendor: {self.vendor_id}, Contract: {COUPON_CONTRACT_ID})", flush=True)

    @staticmethod
    def _timestamp() -> str:
        """현재 시각 문자열 반환"""
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def issue(self) -> None:
        """
        쿠폰 발급 메인 함수

        1. 엑셀에서 쿠폰 정의 읽기
        2. 각 쿠폰에 대해 Coupang API 호출
        3. 결과를 로그로 출력
        """
        timestamp = self._timestamp()
        print(f"[{timestamp}] 쿠폰 발급 작업 시작", flush=True)

        results = []

        try:
            # 1. 엑셀에서 쿠폰 정의 읽기
            coupons = self._fetch_coupons_from_excel()

            if not coupons:
                print(f"[{timestamp}] 발급할 쿠폰이 없습니다", flush=True)
                return

            # 2. 각 쿠폰 발급 처리
            print(f"[{timestamp}] 쿠폰 발급 처리 중: 총 {len(coupons)}개", flush=True)

            for idx, coupon in enumerate(coupons, start=1):
                result = self._issue_single_coupon(idx, coupon)
                results.append(result)

            # 3. 결과 요약 출력
            success_count = sum(1 for r in results if r['status'] == '성공')
            fail_count = len(results) - success_count

            print(f"[{timestamp}] 쿠폰 발급 완료! (성공: {success_count}, 실패: {fail_count})", flush=True)

            # 4. 상세 결과 로깅
            for result in results:
                status = "OK" if result['status'] == '성공' else "FAIL"
                print(f"[{timestamp}] [{status}] {result['coupon_name']}: {result['message']}", flush=True)

        except Exception as e:
            print(f"[{timestamp}] ERROR: 쿠폰 발급 중 오류 발생: {e}", flush=True)
            raise

    def _issue_single_coupon(self, index: int, coupon: Dict[str, Any]) -> Dict[str, Any]:
        """
        단일 쿠폰 발급

        Args:
            index: 쿠폰 순번 (로그용)
            coupon: 쿠폰 정의 딕셔너리
                {
                    'name': '쿠폰이름',
                    'type': '즉시할인' or '다운로드쿠폰',
                    'validity_days': 2,  (유효기간 일수)
                    'discount_type': 'RATE' or 'FIXED_WITH_QUANTITY' or 'PRICE',
                    'discount': 10,  (할인금액/비율, Column E)
                    'issue_count': 100,  (발급개수, Column F, 다운로드쿠폰만 사용)
                }

        Returns:
            발급 결과 딕셔너리
        """
        timestamp = self._timestamp()
        coupon_name = coupon.get('name', f'쿠폰{index}')
        coupon_type = coupon.get('type', '').strip()
        validity_days = coupon.get('validity_days', 1)
        discount_type = coupon.get('discount_type', 'PRICE')
        discount = coupon.get('discount', 0)  # Column E: 할인금액/비율 (필수)
        issue_count = coupon.get('issue_count')  # Column F: 발급개수 (선택적)

        # Validate required fields
        if discount <= 0:
            raise ValueError(f"할인금액/비율이 설정되지 않았습니다: {coupon_name}")

        print(f"[{timestamp}] [{index}] {coupon_name} ({coupon_type}) 발급 중...", flush=True)

        result = {
            'coupon_name': coupon_name,
            'coupon_type': coupon_type,
            'status': '실패',
            'message': '',
            'timestamp': timestamp
        }

        try:
            # 유효 시작일: 오늘 0시
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            start_date = today.strftime('%Y-%m-%d %H:%M:%S')

            # 유효 종료일: 오늘 0시 + validity_days일
            end_date = (today + timedelta(days=validity_days)).strftime('%Y-%m-%d %H:%M:%S')

            if coupon_type == '즉시할인':
                assert self.vendor_id is not None
                api_result = self.api_client.create_instant_coupon(
                    vendor_id=self.vendor_id,
                    contract_id=COUPON_CONTRACT_ID,
                    name=coupon_name,
                    max_discount_price=COUPON_MAX_DISCOUNT,
                    discount=discount,  # Column E: 할인금액/할인율
                    start_at=start_date,
                    end_at=end_date,
                    coupon_type=discount_type
                )

                result['status'] = '성공'
                result['message'] = f"즉시할인쿠폰 생성 완료 (requestedId: {api_result.get('data', {}).get('requestedId', 'N/A')})"

            elif coupon_type == '다운로드쿠폰':
                assert self.user_id is not None
                assert issue_count is not None  # Should have been set by _fetch_coupons_from_excel
                # 다운로드쿠폰 정책 구성
                policy = {
                    "title": coupon_name,
                    "typeOfDiscount": discount_type,
                    "description": f"{coupon_name} ({validity_days}일간 유효)",
                    "discount": discount,  # Column E: 할인금액/할인율
                    "maximumDiscountPrice": COUPON_MAX_DISCOUNT,
                    "maximumPerDaily": issue_count  # Column F: 일 최대 발급개수
                }

                api_result = self.api_client.create_download_coupon(
                    contract_id=COUPON_CONTRACT_ID,
                    title=coupon_name,
                    start_date=start_date,
                    end_date=end_date,
                    user_id=self.user_id,
                    policies=[policy]
                )

                result['status'] = '성공'
                result['message'] = f"다운로드쿠폰 생성 완료 (couponId: {api_result.get('data', {}).get('couponId', 'N/A')})"

            else:
                result['message'] = f"알 수 없는 쿠폰 타입: {coupon_type}"

            print(f"[{timestamp}] [{index}] {result['status']}: {result['message']}", flush=True)

        except Exception as e:
            result['status'] = '실패'
            result['message'] = str(e)
            print(f"[{timestamp}] [{index}] 실패: {e}", flush=True)

        return result

    def _fetch_coupons_from_excel(self) -> List[Dict[str, Any]]:
        """
        엑셀 파일에서 쿠폰 정의 읽기

        엑셀 컬럼 (6개):
        1. 쿠폰이름
        2. 쿠폰타입 (즉시할인 또는 다운로드쿠폰)
        3. 쿠폰유효기간 (일 단위, 예: 2)
        4. 할인방식 (RATE/FIXED_WITH_QUANTITY/PRICE)
        5. 할인금액/비율 (discount value, 할인방식에 따라 의미 다름)
        6. 발급개수 (다운로드쿠폰 전용, 비어있으면 기본값 사용)

        Returns:
            쿠폰 정의 리스트

        Raises:
            FileNotFoundError: 엑셀 파일이 없는 경우
            ValueError: 엑셀 형식이 잘못된 경우
        """
        excel_path = EXCEL_INPUT_FILE  # 이미 Path 객체 (/etc/coupang_coupon_issuer/coupons.xlsx)

        if not excel_path.exists():
            raise FileNotFoundError(f"엑셀 파일이 없습니다: {excel_path}")

        timestamp = self._timestamp()
        print(f"[{timestamp}] 엑셀 파일 읽기: {excel_path}", flush=True)

        try:
            workbook = load_workbook(excel_path, read_only=True)
            sheet = workbook.active

            if sheet is None:
                raise ValueError("엑셀 시트를 찾을 수 없습니다")

            # 헤더 읽기 (첫 번째 행)
            headers = [cell.value for cell in sheet[1]]  # type: ignore

            # 필수 컬럼 체크
            required_columns = ['쿠폰이름', '쿠폰타입', '쿠폰유효기간', '할인방식', '할인금액/비율', '발급개수']

            for col in required_columns:
                if col not in headers:
                    raise ValueError(f"필수 컬럼이 없습니다: {col}")

            # 데이터 행 읽기
            coupons = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # 빈 행 건너뛰기
                if not any(row):
                    continue

                # 컬럼 인덱스 매핑
                col_indices = {header: idx for idx, header in enumerate(headers)}

                # 1. 쿠폰이름: strip만 적용
                coupon_name = str(row[col_indices['쿠폰이름']]).strip()

                # 2. 쿠폰타입: strip + 모든 공백 제거 + '즉시할인' or '다운로드쿠폰' 포함 체크
                coupon_type_raw = str(row[col_indices['쿠폰타입']]).strip()
                coupon_type_normalized = re.sub(r'\s+', '', coupon_type_raw)  # 모든 공백 제거

                if '즉시할인' in coupon_type_normalized:
                    coupon_type = '즉시할인'
                elif '다운로드쿠폰' in coupon_type_normalized:
                    coupon_type = '다운로드쿠폰'
                else:
                    raise ValueError(
                        f"행 {row_idx}: 잘못된 쿠폰 타입 '{coupon_type_raw}' (즉시할인 또는 다운로드쿠폰만 가능)"
                    )

                # 3. 쿠폰유효기간: 숫자만 추출 -> float -> int
                validity_days_raw = str(row[col_indices['쿠폰유효기간']])
                validity_days_digits = re.sub(r'[^\d.]', '', validity_days_raw)  # 숫자와 소수점만 남김
                try:
                    validity_days = int(float(validity_days_digits))
                    if validity_days <= 0:
                        raise ValueError(f"행 {row_idx}: 쿠폰유효기간은 1 이상이어야 합니다")
                except (ValueError, TypeError):
                    raise ValueError(f"행 {row_idx}: 쿠폰유효기간은 숫자여야 합니다 (현재값: {validity_days_raw})")

                # 4. 할인방식: upper case + '-'를 '_'로 변환 + 포함 체크
                discount_type_raw = str(row[col_indices['할인방식']]).strip()
                discount_type_normalized = discount_type_raw.upper().replace('-', '_')

                if 'RATE' in discount_type_normalized:
                    discount_type = 'RATE'
                elif 'FIXED_WITH_QUANTITY' in discount_type_normalized or 'FIXED WITH QUANTITY' in discount_type_normalized.replace('_', ' '):
                    discount_type = 'FIXED_WITH_QUANTITY'
                elif 'PRICE' in discount_type_normalized:
                    discount_type = 'PRICE'
                else:
                    raise ValueError(
                        f"행 {row_idx}: 잘못된 할인방식 '{discount_type_raw}' (RATE/FIXED_WITH_QUANTITY/PRICE만 가능)"
                    )

                # 5. 할인금액/비율: 숫자만 추출 -> float -> int
                discount_raw = str(row[col_indices['할인금액/비율']])
                discount_digits = re.sub(r'[^\d.]', '', discount_raw)  # 숫자와 소수점만 남김
                try:
                    discount = int(float(discount_digits)) if discount_digits else 0
                except (ValueError, TypeError):
                    raise ValueError(f"행 {row_idx}: 할인금액/비율은 숫자여야 합니다 (현재값: {discount_raw})")

                if discount <= 0:
                    raise ValueError(f"행 {row_idx}: 할인금액/비율은 0보다 커야 합니다")

                # 6. 발급개수: 선택적 (쿠폰 타입에 따라 처리)
                issue_count_raw = str(row[col_indices['발급개수']]).strip()
                issue_count = None

                # 즉시할인: 발급개수 무시 (API에서 사용 안함)
                if coupon_type == '즉시할인':
                    issue_count = None  # Not used in API

                # 다운로드쿠폰: 발급개수 필요 (비어있으면 기본값)
                elif coupon_type == '다운로드쿠폰':
                    if issue_count_raw and issue_count_raw != 'None':
                        issue_count_digits = re.sub(r'[^\d.]', '', issue_count_raw)
                        try:
                            issue_count = int(float(issue_count_digits)) if issue_count_digits else COUPON_DEFAULT_ISSUE_COUNT
                            if issue_count < 1:
                                raise ValueError(f"행 {row_idx}: 발급개수는 1 이상이어야 합니다 (현재: {issue_count})")
                        except (ValueError, TypeError):
                            raise ValueError(f"행 {row_idx}: 발급개수는 숫자여야 합니다 (현재값: {issue_count_raw})")
                    else:
                        issue_count = COUPON_DEFAULT_ISSUE_COUNT  # Default value

                # 7. 할인방식별 추가 검증 (Column E '할인금액/비율' 기준)
                if discount_type == 'RATE':
                    # 정률할인: 1~99% 범위 체크
                    if not (1 <= discount <= 99):
                        raise ValueError(f"행 {row_idx}: RATE 할인율은 1~99 사이여야 합니다 (현재: {discount})")
                elif discount_type == 'PRICE':
                    # 정액할인: 10원 단위 및 최소 10원 체크
                    if discount < 10:
                        raise ValueError(f"행 {row_idx}: PRICE 할인금액은 최소 10원 이상이어야 합니다 (현재: {discount})")
                    if discount % 10 != 0:
                        raise ValueError(f"행 {row_idx}: PRICE 할인금액은 10원 단위여야 합니다 (현재: {discount})")
                elif discount_type == 'FIXED_WITH_QUANTITY':
                    # 수량할인: 1 이상 체크
                    if discount < 1:
                        raise ValueError(f"행 {row_idx}: FIXED_WITH_QUANTITY 할인은 1 이상이어야 합니다 (현재: {discount})")

                coupon = {
                    'name': coupon_name,
                    'type': coupon_type,
                    'validity_days': validity_days,
                    'discount_type': discount_type,
                    'discount': discount,  # NEW: from column E
                    'issue_count': issue_count,  # From column F (None for instant coupons)
                }

                coupons.append(coupon)

            workbook.close()

            print(f"[{timestamp}] 쿠폰 {len(coupons)}개 읽기 완료", flush=True)
            return coupons

        except Exception as e:
            print(f"[{timestamp}] ERROR: 엑셀 파일 읽기 실패: {e}", flush=True)
            raise
