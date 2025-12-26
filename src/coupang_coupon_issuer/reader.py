"""엑셀 파일에서 쿠폰 정의를 읽고 검증하는 모듈"""

import re
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook

from .config import COUPON_DEFAULT_ISSUE_COUNT

# 할인방식 한글-영어 매핑
DISCOUNT_TYPE_KR_TO_EN = {
    '정률할인': 'RATE',
    '수량별 정액할인': 'FIXED_WITH_QUANTITY',
    '정액할인': 'PRICE',
}

DISCOUNT_TYPE_EN_TO_KR = {
    'RATE': '정률할인',
    'FIXED_WITH_QUANTITY': '수량별 정액할인',
    'PRICE': '정액할인',
}


def fetch_coupons_from_excel(excel_path: Path) -> List[Dict[str, Any]]:
    """
    엑셀 파일에서 쿠폰 정의 읽기 및 검증

    엑셀 컬럼 (7개):
    1. 쿠폰이름
    2. 쿠폰타입 (즉시할인쿠폰 또는 다운로드쿠폰)
    3. 쿠폰유효기간 (일 단위, 예: 2)
    4. 할인방식 (정률할인/수량별 정액할인/정액할인)
    5. 할인금액/비율 (discount value, 할인방식에 따라 의미 다름)
    6. 발급개수 (다운로드쿠폰 전용, 비어있으면 기본값 사용)
    7. 옵션ID (쉼표로 구분된 vendor item ID 리스트, 필수)

    Returns:
        쿠폰 정의 리스트
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {excel_path}")

    try:
        workbook = load_workbook(excel_path, read_only=True)
        sheet = workbook.active

        if sheet is None:
            raise ValueError("엑셀 시트를 찾을 수 없습니다")

        # 헤더 읽기 (첫 번째 행)
        headers = [cell.value for cell in sheet[1]]  # type: ignore

        # 필수 컬럼 체크 (9개)
        required_columns = [
            '쿠폰이름', '쿠폰타입', '쿠폰유효기간', '할인방식', 
            '할인금액/비율', '최소구매금액', '최대할인금액',
            '발급개수', '옵션ID'
        ]

        for col in required_columns:
            # 옵션ID는 "옵션 ID" (공백 포함)도 허용 (ADR 002 입력 정규화)
            if col == '옵션ID' and '옵션 ID' in headers:
                continue
            if col not in headers:
                raise ValueError(f"필수 컬럼이 없습니다: {col}")

        # 데이터 행 읽기
        coupons = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # 빈 행 건너뛰기
            if not any(row):
                continue

            # 컬럼 인덱스 매핑 (옵션ID는 "옵션 ID"도 허용)
            col_indices = {header: idx for idx, header in enumerate(headers)}
            if '옵션ID' not in col_indices and '옵션 ID' in col_indices:
                col_indices['옵션ID'] = col_indices['옵션 ID']

            # 1. 쿠폰이름: strip만 적용
            coupon_name = str(row[col_indices['쿠폰이름']]).strip()

            # 2. 쿠폰타입: strip + 모든 공백 제거 + '즉시할인쿠폰' or '다운로드쿠폰' 포함 체크
            coupon_type_raw = str(row[col_indices['쿠폰타입']]).strip()
            coupon_type_normalized = re.sub(r'\s+', '', coupon_type_raw)  # 모든 공백 제거

            if '즉시할인' in coupon_type_normalized:
                coupon_type = '즉시할인쿠폰'  # 하위 호환: "즉시할인" 입력도 허용
            elif '다운로드쿠폰' in coupon_type_normalized:
                coupon_type = '다운로드쿠폰'
            else:
                raise ValueError(
                    f"행 {row_idx}: 잘못된 쿠폰 타입 '{coupon_type_raw}' (즉시할인쿠폰 또는 다운로드쿠폰만 가능)"
                )

            # 3. 쿠폰유효기간: 숫자만 추출 -> float -> int
            validity_days_raw = str(row[col_indices['쿠폰유효기간']])
            validity_days_digits = re.sub(r'[^\d.]', '', validity_days_raw)  # 숫자와 소수점만 남김
            try:
                validity_days = int(float(validity_days_digits))
                if validity_days <= 0:
                    raise ValueError(f"행 {row_idx}: 쿠폰유효기간은 1 이상이어야 합니다")
            except (ValueError, TypeError):
                raise ValueError(f"행 {row_idx}: 쿠폰유효기간은 숫자여야 합니다 (현재값: {validity_days_raw})")

            # 4. 할인방식: 한글 입력 지원
            discount_type_raw = str(row[col_indices['할인방식']]).strip()
            
            # 한글 매핑 시도
            if discount_type_raw in DISCOUNT_TYPE_KR_TO_EN:
                discount_type = DISCOUNT_TYPE_KR_TO_EN[discount_type_raw]
            else:
                raise ValueError(
                    f"행 {row_idx}: 잘못된 할인방식 '{discount_type_raw}' (정률할인/수량별 정액할인/정액할인만 가능)"
                )

            # 5. 할인금액/비율: 숫자만 추출 -> float -> int
            discount_raw = str(row[col_indices['할인금액/비율']])
            discount_digits = re.sub(r'[^\d.]', '', discount_raw)  # 숫자와 소수점만 남김
            try:
                discount = int(float(discount_digits)) if discount_digits else 0
            except (ValueError, TypeError):
                raise ValueError(f"행 {row_idx}: 할인금액/비율은 숫자여야 합니다 (현재값: {discount_raw})")

            if discount <= 0:
                raise ValueError(f"행 {row_idx}: 할인금액/비율은 0보다 커야 합니다")

            # 6. 최소구매금액 (Column F): 다운로드쿠폰 전용, 최소 구매 조건 (선택적, 기본값 1)
            min_purchase_raw = str(row[col_indices['최소구매금액']]).strip()
            min_purchase_price = None

            if coupon_type == '다운로드쿠폰':
                # 다운로드쿠폰: 사용자 입력 또는 기본값 1원
                if min_purchase_raw and min_purchase_raw != 'None':
                    min_purchase_digits = re.sub(r'[^\d.]', '', min_purchase_raw)
                    try:
                        min_purchase_price = int(float(min_purchase_digits)) if min_purchase_digits else 1
                    except (ValueError, TypeError):
                        raise ValueError(f"행 {row_idx}: 최소구매금액은 숫자여야 합니다 (현재값: {min_purchase_raw})")
                    
                    if min_purchase_price < 1:
                        raise ValueError(f"행 {row_idx}: 최소구매금액은 1원 이상이어야 합니다 (현재: {min_purchase_price})")
                else:
                    min_purchase_price = 10  # 기본값 (API 최소값: 10원)
            elif coupon_type == '즉시할인쿠폰':
                # 즉시할인쿠폰: 사용 안함
                min_purchase_price = None

            # 7. 최대할인금액 (Column G): 정률할인 시 최대 할인 금액 (필수, 양의 정수)
            max_discount_raw = str(row[col_indices['최대할인금액']]).strip()
            max_discount_digits = re.sub(r'[^\d.]', '', max_discount_raw)
            try:
                max_discount_price = int(float(max_discount_digits)) if max_discount_digits else 0
            except (ValueError, TypeError):
                raise ValueError(f"행 {row_idx}: 최대할인금액은 숫자여야 합니다 (현재값: {max_discount_raw})")

            if max_discount_price <= 0:
                raise ValueError(f"행 {row_idx}: 최대할인금액은 0보다 커야 합니다")

            # 8. 발급개수: 선택적 (쿠폰 타입에 따라 처리)
            issue_count_raw = str(row[col_indices['발급개수']]).strip()
            issue_count = None

            # 즉시할인쿠폰: 발급개수 무시 (API에서 사용 안함)
            if coupon_type == '즉시할인쿠폰':
                issue_count = None  # Not used in API

            # 다운로드쿠폰: 발급개수 필요 (비어있으면 기본값)
            elif coupon_type == '다운로드쿠폰':
                if issue_count_raw and issue_count_raw != 'None':
                    issue_count_digits = re.sub(r'[^\d.]', '', issue_count_raw)
                    try:
                        issue_count = int(float(issue_count_digits)) if issue_count_digits else COUPON_DEFAULT_ISSUE_COUNT
                    except (ValueError, TypeError):
                        raise ValueError(f"행 {row_idx}: 발급개수는 숫자여야 합니다 (현재값: {issue_count_raw})")

                    if issue_count < 1:
                        raise ValueError(f"행 {row_idx}: 발급개수는 1 이상이어야 합니다 (현재: {issue_count})")
                else:
                    issue_count = COUPON_DEFAULT_ISSUE_COUNT  # Default value

            # 7. 쿠폰 타입 + 할인방식별 검증 (Column E '할인금액/비율' 기준)
            # ADR 017: 쿠폰 타입별로 검증 규칙이 다름
            if coupon_type == '다운로드쿠폰':
                # 다운로드 쿠폰 검증 규칙
                if discount_type == 'RATE':
                    # 정률할인: 1~99% 범위 체크 (100% 불가)
                    if not (1 <= discount <= 99):
                        raise ValueError(f"행 {row_idx}: 다운로드쿠폰 정률할인은 1~99 사이여야 합니다 (현재: {discount})")
                elif discount_type == 'PRICE':
                    # 정액할인: 10원 단위 및 최소 10원 체크
                    if discount < 10:
                        raise ValueError(f"행 {row_idx}: 다운로드쿠폰 정액할인은 최소 10원 이상이어야 합니다 (현재: {discount})")
                    if discount % 10 != 0:
                        raise ValueError(f"행 {row_idx}: 다운로드쿠폰 정액할인은 10원 단위여야 합니다 (현재: {discount})")
            elif coupon_type == '즉시할인쿠폰':
                # 즉시할인쿠폰 검증 규칙
                if discount_type == 'RATE':
                    # 정률할인: 1~100% 범위 체크 (100% 허용)
                    if not (1 <= discount <= 100):
                        raise ValueError(f"행 {row_idx}: 즉시할인쿠폰 정률할인은 1~100 사이여야 합니다 (현재: {discount})")
                elif discount_type == 'PRICE':
                    # 정액할인: 1원 이상 체크 (10원 단위 제약 없음)
                    if discount < 1:
                        raise ValueError(f"행 {row_idx}: 즉시할인쿠폰 정액할인은 1원 이상이어야 합니다 (현재: {discount})")
                elif discount_type == 'FIXED_WITH_QUANTITY':
                    # 수량할인: 1 이상 체크
                    if discount < 1:
                        raise ValueError(f"행 {row_idx}: 즉시할인쿠폰 수량별 정액할인은 1 이상이어야 합니다 (현재: {discount})")

            # 9. 옵션ID (Column I): 쉼표로 구분된 vendor item ID 리스트 (필수)
            vendor_items_raw = str(row[col_indices['옵션ID']]).strip()

            if not vendor_items_raw or vendor_items_raw == 'None':
                raise ValueError(f"행 {row_idx}: 옵션ID는 필수 입력입니다")

            # 쉼표로 분리 + strip + int 변환
            try:
                vendor_items = [
                    int(item.strip())
                    for item in vendor_items_raw.split(',')
                    if item.strip()
                ]
            except (ValueError, TypeError):
                raise ValueError(f"행 {row_idx}: 옵션ID는 숫자만 입력 가능합니다 (현재값: {vendor_items_raw})")

            if not vendor_items:
                raise ValueError(f"행 {row_idx}: 옵션ID가 비어있습니다")

            # Coupang API 제한 검증
            if coupon_type == '즉시할인쿠폰' and len(vendor_items) > 10000:
                raise ValueError(f"행 {row_idx}: 즉시할인쿠폰은 최대 10,000개의 옵션ID만 지원합니다 (현재: {len(vendor_items)}개)")
            elif coupon_type == '다운로드쿠폰' and len(vendor_items) > 100:
                raise ValueError(f"행 {row_idx}: 다운로드쿠폰은 최대 100개의 옵션ID만 지원합니다 (현재: {len(vendor_items)}개)")

            # 양의 정수 검증
            for vid in vendor_items:
                if vid <= 0:
                    raise ValueError(f"행 {row_idx}: 옵션ID는 양의 정수여야 합니다 (현재: {vid})")

            coupon = {
                'name': coupon_name,
                'type': coupon_type,
                'validity_days': validity_days,
                'discount_type': discount_type,
                'discount': discount,  # Column E: 할인금액/비율
                'min_purchase_price': min_purchase_price,  # Column F: 최소구매금액 (다운로드쿠폰 전용)
                'max_discount_price': max_discount_price,  # Column G: 최대할인금액
                'issue_count': issue_count,  # Column H: 발급개수 (None for instant coupons)
                'vendor_items': vendor_items,  # Column I: 옵션ID 리스트
            }

            coupons.append(coupon)

        workbook.close()
        return coupons

    except Exception as e:
        raise ValueError(f"엑셀 파일 읽기 실패: {e}")
